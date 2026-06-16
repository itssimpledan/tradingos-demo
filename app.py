"""
app.py — Flask backend for the Options Trading System
Run: python app.py  →  http://localhost:5000
"""

import csv
import io
import json
import math
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from flask import Flask, jsonify, request, render_template, abort, Response
import yfinance as yf
import logging
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
logging.getLogger("urllib3").setLevel(logging.CRITICAL)
logging.getLogger("peewee").setLevel(logging.CRITICAL)

from database import (
    init_db,
    # platforms
    list_platforms, add_platform, delete_platform,
    # equities
    add_equity, list_equities, get_equity, update_equity,
    update_equity_price, delete_equity, bulk_insert_equities, equities_summary,
    # purchases
    list_purchases, add_purchase, delete_purchase,
    # crypto
    add_crypto, list_crypto, get_crypto, update_crypto,
    update_crypto_price, delete_crypto, crypto_summary,
    # cash
    list_cash, add_cash, update_cash, delete_cash,
    # options trades
    add_trade, close_trade, update_trade, get_trade, list_trades, delete_trade,
    # watchlist
    add_to_watchlist, list_watchlist, remove_from_watchlist,
    # backtests
    save_backtest, list_backtests, get_backtest,
    # dashboard
    portfolio_summary,
    # net worth
    list_networth_items, add_networth_item, update_networth_item, delete_networth_item,
    # monthly snapshots
    save_monthly_snapshot, list_monthly_snapshots, get_monthly_snapshot, delete_monthly_snapshot,
)
from screener    import screen_tickers, get_universe, UNIVERSES
from recommender import recommend
from analyser    import analyse
from backtester  import run_backtest

app = Flask(__name__)

# ─── Demo mode (public showcase deployment) ──
# Set DEMO_MODE=1 in the environment to show the demo banner and cap the
# AI assistant so a visitor can't run up the owner's Anthropic bill.
DEMO_MODE          = os.environ.get("DEMO_MODE", "0") == "1"
DEMO_ASSISTANT_CAP = int(os.environ.get("DEMO_ASSISTANT_CAP", "5"))   # msgs per visitor per day
_assistant_usage    = defaultdict(list)   # ip -> [timestamps]

def _assistant_quota_ok(ip: str) -> bool:
    """True if this IP still has demo assistant messages left today."""
    now = time.time()
    window = 24 * 3600
    hits = [t for t in _assistant_usage[ip] if now - t < window]
    _assistant_usage[ip] = hits
    return len(hits) < DEMO_ASSISTANT_CAP

def _assistant_quota_hit(ip: str):
    _assistant_usage[ip].append(time.time())

# Free hosting tiers (e.g. PythonAnywhere free) block outbound calls to
# Yahoo Finance, so yfinance would fail on every route. In demo mode we
# swap it for a synthetic-but-realistic data shim instead of letting the
# whole app 502. See demo_market_data.py for details.
if DEMO_MODE:
    from demo_market_data import patch_yfinance
    patch_yfinance()

# ─── openpyxl (optional — for xlsx import) ──
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# ─── pdf_report (optional — for PDF export) ──
try:
    from pdf_report import build_portfolio_pdf
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


# ─────────────────────────────────────────────
# Startup — initialise DB + pre-warm FX cache
# ─────────────────────────────────────────────

init_db()

# _prewarm_fx is defined below after all helpers — thread is started there


@app.route("/")
def index():
    return render_template("index.html", demo_mode=DEMO_MODE, demo_cap=DEMO_ASSISTANT_CAP)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

_fx_cache: dict = {}   # { "HKDSGD": (rate, timestamp), ... }
_FX_CACHE_TTL = 4 * 3600  # 4-hour cache — FX rates change slowly

# Fallback rates (used when live fetch fails)
_FX_FALLBACKS = {
    "USDSGD": 1.35, "HKDSGD": 0.173, "GBPSGD": 1.71,
    "EURSGD": 1.47, "JPYSGD": 0.0090, "AUDSGD": 0.88,
    "CADSGD": 1.00, "CNYSGD": 0.187, "USDGBP": 0.79,
    "USDEUR": 0.92, "USDHKD": 7.82,
}

def get_fx_rate(from_ccy: str = "USD", to_ccy: str = "SGD") -> float:
    """Fetch live FX rate via yfinance with a 4-hour in-memory cache."""
    if from_ccy == to_ccy:
        return 1.0
    key = f"{from_ccy}{to_ccy}"
    cached = _fx_cache.get(key)
    if cached:
        rate, ts = cached
        if (datetime.now() - ts).total_seconds() < _FX_CACHE_TTL:
            return rate
    try:
        hist = yf.Ticker(f"{key}=X").history(period="1d")
        if not hist.empty:
            rate = round(float(hist["Close"].iloc[-1]), 4)
            _fx_cache[key] = (rate, datetime.now())
            return rate
    except Exception:
        pass
    return _FX_FALLBACKS.get(key, 1.0)


def _batch_fx_rates(pairs: list[str]) -> None:
    """Fetch multiple FX pairs in ONE yf.download() call and populate cache.
    pairs: list of 'FROMTO' strings e.g. ['USDSGD', 'HKDSGD'].
    Already-cached (and still fresh) pairs are skipped.
    """
    now = datetime.now()
    needed = []
    for key in pairs:
        cached = _fx_cache.get(key)
        if not cached or (now - cached[1]).total_seconds() >= _FX_CACHE_TTL:
            needed.append(key)
    if not needed:
        return
    tickers = [f"{k}=X" for k in needed]
    try:
        import pandas as pd
        df = yf.download(tickers, period="1d", auto_adjust=True,
                         progress=False, threads=True)
        close = df["Close"] if "Close" in df else df
        if isinstance(close, pd.Series):
            # single ticker returned as Series
            key  = needed[0]
            vals = close.dropna()
            if not vals.empty:
                _fx_cache[key] = (round(float(vals.iloc[-1]), 4), now)
        else:
            for key in needed:
                col = f"{key}=X"
                if col in close.columns:
                    vals = close[col].dropna()
                    if not vals.empty:
                        _fx_cache[key] = (round(float(vals.iloc[-1]), 4), now)
    except Exception:
        pass  # fall back to individual cached-fallback lookups


def _to_sgd_rate(ccy: str, usdsgd: float) -> float:
    """Return conversion rate: 1 unit of ccy → SGD."""
    ccy = (ccy or "USD").upper()
    if ccy == "SGD":
        return 1.0
    if ccy == "USD":
        return usdsgd
    # Use cache-first lookup (populated by _batch_fx_rates before enrichment)
    key = f"{ccy}SGD"
    cached = _fx_cache.get(key)
    if cached and (datetime.now() - cached[1]).total_seconds() < _FX_CACHE_TTL:
        return cached[0]
    if key in _FX_FALLBACKS:
        return get_fx_rate(ccy, "SGD")
    # Route via USD: ccy→USD→SGD
    to_usd = get_fx_rate(ccy, "USD")
    return round(to_usd * usdsgd, 6)


def _prewarm_fx() -> None:
    """Background thread: fetch common FX pairs so first API call is instant."""
    common = ["USDSGD", "HKDSGD", "GBPSGD", "EURSGD", "AUDSGD", "CADSGD", "JPYSGD"]
    # Also fetch any currencies currently in the DB
    try:
        from database import list_equities as _le
        rows = _le()
        db_ccys = {(r.get("currency") or "USD").upper() for r in rows if r.get("currency", "USD") != "SGD"}
        for c in db_ccys:
            if c != "USD":
                common.append(f"{c}SGD")
        common.append("USDSGD")  # always need this one
    except Exception:
        pass
    _batch_fx_rates(list(dict.fromkeys(common)))  # deduplicated


# Start FX pre-warm now that the function is defined
import threading as _threading
_threading.Thread(target=_prewarm_fx, daemon=True, name="fx-prewarm").start()


def _enrich_equity(eq: dict, fx: float = None) -> dict:
    """Add computed P&L columns (local ccy and SGD) to a holding row."""
    if fx is None:
        fx = get_fx_rate("USD", "SGD")
    cost   = eq.get("cost_price") or 0
    price  = eq.get("current_price") or cost
    shares = eq.get("shares") or 0
    ccy    = (eq.get("currency") or "USD").upper()

    mkt_local  = round(shares * price, 2)
    cost_local = round(shares * cost, 2)
    pnl_local  = round(mkt_local - cost_local, 2)
    pnl_pct    = round(pnl_local / cost_local * 100, 2) if cost_local else 0

    # Always convert to SGD using the correct rate for the holding's currency
    to_sgd   = _to_sgd_rate(ccy, fx)
    mkt_sgd  = round(mkt_local  * to_sgd, 2)
    pnl_sgd  = round(pnl_local  * to_sgd, 2)
    cost_sgd = round(cost_local * to_sgd, 2)

    return {**eq,
            "market_value_usd"   : mkt_local,   # local currency value
            "market_value_sgd"   : mkt_sgd,
            "cost_basis_usd"     : cost_local,
            "cost_basis_sgd"     : cost_sgd,
            "unrealized_pnl_usd" : pnl_local,
            "unrealized_pnl_sgd" : pnl_sgd,
            "unrealized_pnl_pct" : pnl_pct,
            "fx_rate"            : to_sgd,
            "currency"           : ccy}


def _enrich_crypto(cr: dict, fx: float = None) -> dict:
    """Add computed P&L columns to a crypto row.
    cost_price is stored in SGD; current_price is in USD (from yfinance).
    """
    if fx is None:
        fx = get_fx_rate("USD", "SGD")
    cost_sgd  = cr.get("cost_price") or 0        # user-entered SGD cost per unit
    price_usd = cr.get("current_price") or 0
    qty       = cr.get("quantity") or 0
    mkt_usd   = round(qty * price_usd, 2)
    mkt_sgd   = round(mkt_usd * fx, 2)
    cost_basis_sgd = round(qty * cost_sgd, 2)
    pnl_sgd   = round(mkt_sgd - cost_basis_sgd, 2)
    pnl_pct   = round(pnl_sgd / cost_basis_sgd * 100, 2) if cost_basis_sgd else 0
    return {**cr,
            "market_value_usd"   : mkt_usd,
            "market_value_sgd"   : mkt_sgd,
            "cost_basis_sgd"     : cost_basis_sgd,
            "unrealized_pnl_usd" : round(pnl_sgd / fx, 2) if fx else 0,
            "unrealized_pnl_sgd" : pnl_sgd,
            "unrealized_pnl_pct" : pnl_pct,
            "fx_rate"            : fx}


def _read_uploaded_file(f) -> list[dict]:
    """Parse a CSV or XLSX file upload into a list of dicts."""
    filename = f.filename.lower()
    if filename.endswith(".xlsx"):
        if not XLSX_AVAILABLE:
            raise RuntimeError("openpyxl not installed — run: pip install openpyxl")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append({headers[i]: (str(v).strip() if v is not None else "")
                             for i, v in enumerate(row) if i < len(headers)})
        return rows
    else:
        content = f.stream.read().decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(content)))


def _yf_silence():
    """Context manager: suppress stdout, stderr, and yfinance/urllib logging."""
    import sys, io, logging, contextlib
    @contextlib.contextmanager
    def _ctx():
        old_out, old_err = sys.stdout, sys.stderr
        sys.stdout = sys.stderr = io.StringIO()
        loggers = ['yfinance', 'urllib3', 'urllib3.connectionpool', 'requests']
        saved = {}
        for name in loggers:
            lg = logging.getLogger(name)
            saved[name] = lg.level
            lg.setLevel(logging.CRITICAL)
        try:
            yield
        finally:
            sys.stdout, sys.stderr = old_out, old_err
            for name, lvl in saved.items():
                logging.getLogger(name).setLevel(lvl)
    return _ctx()


def _fetch_info(ticker: str) -> dict:
    """Fetch company/sector/industry from yfinance. Silences noisy 404s for ETFs/HK stocks."""
    try:
        with _yf_silence():
            info = yf.Ticker(ticker).info or {}
        return {
            "company" : info.get("longName"),
            "sector"  : info.get("sector"),
            "industry": info.get("industry"),
        }
    except Exception:
        return {}


def _safe_yf_info(ticker: str) -> dict:
    """Call yf.Ticker.info with all noise suppressed (stdout, stderr, logging)."""
    try:
        with _yf_silence():
            return yf.Ticker(ticker).info or {}
    except Exception:
        return {}

# ─────────────────────────────────────────────
# FX rate endpoint
# ─────────────────────────────────────────────

@app.route("/api/fx/<pair>")
def fx_rate(pair):
    pair = pair.upper()
    from_ccy = pair[:3]
    to_ccy   = pair[3:]
    rate = get_fx_rate(from_ccy, to_ccy)
    return jsonify({"pair": pair, "rate": rate, "from": from_ccy, "to": to_ccy})





# ═══════════════════════════════════════════════
# SCREENER
# ═══════════════════════════════════════════════

@app.route("/api/screen", methods=["POST"])
def screen():
    body     = request.get_json() or {}
    tickers  = body.get("tickers", [])
    universe = body.get("universe")
    criteria = body.get("criteria", {})
    if universe:
        tickers = get_universe(universe)
    if not tickers:
        return jsonify({"error": "Provide 'tickers' list or 'universe' name"}), 400
    results = screen_tickers(tickers, criteria)
    return jsonify({"results": results, "count": len([r for r in results if r.get("passed")])})


@app.route("/api/universes")
def universes():
    return jsonify(UNIVERSES)


# ═══════════════════════════════════════════════
# THEME SCREENER
# ═══════════════════════════════════════════════

# Curated large-cap US ticker lists per theme (no API key needed)
THEME_TICKERS = {
    "AI Infrastructure": [
        {"ticker":"NVDA",  "company":"NVIDIA Corporation",           "role":"Designs GPUs that power AI training and inference at scale"},
        {"ticker":"AMD",   "company":"Advanced Micro Devices",        "role":"GPU and CPU chips competing with NVIDIA for AI compute"},
        {"ticker":"INTC",  "company":"Intel Corporation",             "role":"Data-centre CPUs and Gaudi AI accelerators"},
        {"ticker":"AVGO",  "company":"Broadcom Inc.",                 "role":"Custom AI ASICs and networking chips for hyperscalers"},
        {"ticker":"MRVL",  "company":"Marvell Technology",            "role":"Custom silicon and data-centre networking for AI clusters"},
        {"ticker":"MSFT",  "company":"Microsoft Corporation",         "role":"Azure cloud platform and largest investor in OpenAI"},
        {"ticker":"AMZN",  "company":"Amazon.com Inc.",               "role":"AWS cloud infrastructure and Trainium/Inferentia AI chips"},
        {"ticker":"GOOGL", "company":"Alphabet Inc.",                 "role":"Google Cloud, TPU AI chips and DeepMind research"},
        {"ticker":"META",  "company":"Meta Platforms Inc.",           "role":"Builds massive AI compute clusters for recommendation and LLMs"},
        {"ticker":"ORCL",  "company":"Oracle Corporation",            "role":"Fastest-growing AI cloud GPU cluster rental platform"},
        {"ticker":"CRM",   "company":"Salesforce Inc.",               "role":"AI-powered CRM and Einstein AI platform"},
        {"ticker":"ANET",  "company":"Arista Networks",               "role":"High-speed Ethernet switching for AI data-centre networking"},
        {"ticker":"TSM",   "company":"Taiwan Semiconductor (US ADR)", "role":"World's leading foundry manufacturing AI chips"},
        {"ticker":"QCOM",  "company":"Qualcomm Inc.",                 "role":"Edge AI processors for mobile and IoT devices"},
        {"ticker":"SMCI",  "company":"Super Micro Computer",          "role":"AI server systems integrating NVIDIA GPUs at high density"},
        {"ticker":"VST",   "company":"Vistra Corp.",                  "role":"Power utility supplying electricity to AI data centres"},
        {"ticker":"CEG",   "company":"Constellation Energy",          "role":"Nuclear power provider for carbon-free AI data-centre energy"},
        {"ticker":"EQIX",  "company":"Equinix Inc.",                  "role":"Global data-centre colocation for AI workloads"},
        {"ticker":"DLR",   "company":"Digital Realty Trust",          "role":"Data-centre REIT powering hyperscaler AI infrastructure"},
        {"ticker":"DELL",  "company":"Dell Technologies",             "role":"AI server and storage solutions powered by NVIDIA GPUs"},
        {"ticker":"HPE",   "company":"Hewlett Packard Enterprise",    "role":"AI supercomputing systems including Cray and HPE Proliant"},
        {"ticker":"CDNS",  "company":"Cadence Design Systems",        "role":"EDA software used to design every AI chip"},
        {"ticker":"SNPS",  "company":"Synopsys Inc.",                 "role":"EDA and semiconductor IP critical to AI chip development"},
        {"ticker":"APH",   "company":"Amphenol Corporation",          "role":"High-speed connectors used inside AI server racks"},
        {"ticker":"GLW",   "company":"Corning Inc.",                  "role":"Optical fibre for the high-bandwidth AI network backbone"},
    ],
    "AI Application Layer": [
        {"ticker":"MSFT",  "company":"Microsoft Corporation",         "role":"Copilot AI integrated across Office 365 and Azure for enterprise"},
        {"ticker":"GOOGL", "company":"Alphabet Inc.",                 "role":"Gemini AI embedded in Search, Workspace and Google Cloud"},
        {"ticker":"META",  "company":"Meta Platforms Inc.",           "role":"AI-driven ad targeting and Llama open-source models"},
        {"ticker":"CRM",   "company":"Salesforce Inc.",               "role":"Einstein AI and Agentforce automating CRM workflows"},
        {"ticker":"NOW",   "company":"ServiceNow Inc.",               "role":"Now Assist AI agents automating enterprise IT workflows"},
        {"ticker":"ADBE",  "company":"Adobe Inc.",                    "role":"Firefly generative AI monetised across Creative Cloud"},
        {"ticker":"ORCL",  "company":"Oracle Corporation",            "role":"Fusion Cloud AI automating ERP and HR at enterprises"},
        {"ticker":"SAP",   "company":"SAP SE (US ADR)",               "role":"Joule AI copilot embedded across SAP enterprise software"},
        {"ticker":"INTU",  "company":"Intuit Inc.",                   "role":"AI-powered tax, accounting and small-business financial tools"},
        {"ticker":"VEEV",  "company":"Veeva Systems Inc.",            "role":"AI tools for life-sciences CRM and clinical data management"},
        {"ticker":"WDAY",  "company":"Workday Inc.",                  "role":"AI agents automating HR and finance workflows in enterprise"},
        {"ticker":"PLTR",  "company":"Palantir Technologies",         "role":"AIP platform deploying AI decision-making in enterprise and defence"},
        {"ticker":"AI",    "company":"C3.ai Inc.",                    "role":"Enterprise AI application suite across energy, defence and finance"},
        {"ticker":"SMAR",  "company":"Smartsheet Inc.",               "role":"AI-enhanced work management automating project coordination"},
        {"ticker":"HUBS",  "company":"HubSpot Inc.",                  "role":"AI-powered marketing and sales CRM for SMBs"},
        {"ticker":"DDOG",  "company":"Datadog Inc.",                  "role":"AI observability platform monitoring LLM application performance"},
        {"ticker":"SNOW",  "company":"Snowflake Inc.",                "role":"Data cloud platform enabling AI/ML on enterprise data at scale"},
        {"ticker":"MDB",   "company":"MongoDB Inc.",                  "role":"Document database underpinning AI application data layers"},
        {"ticker":"GTLB",  "company":"GitLab Inc.",                   "role":"AI-powered DevSecOps platform accelerating software development"},
        {"ticker":"PATH",  "company":"UiPath Inc.",                   "role":"AI-powered RPA automating repetitive enterprise processes"},
        {"ticker":"ZS",    "company":"Zscaler Inc.",                  "role":"AI-driven zero-trust security platform for cloud workloads"},
        {"ticker":"CRWD",  "company":"CrowdStrike Holdings",          "role":"Charlotte AI threat detection across endpoint and cloud"},
        {"ticker":"TEAM",  "company":"Atlassian Corporation",         "role":"Rovo AI agents automating team collaboration across Jira and Confluence"},
        {"ticker":"AMZN",  "company":"Amazon.com Inc.",               "role":"Alexa+ and AWS Bedrock monetising AI in retail and cloud"},
        {"ticker":"IBM",   "company":"IBM Corporation",               "role":"watsonx enterprise AI platform for regulated industries"},
    ],
    "Nuclear & Grid Renaissance": [
        {"ticker":"CEG",   "company":"Constellation Energy",          "role":"Largest US nuclear operator; restarted Three Mile Island for Microsoft"},
        {"ticker":"VST",   "company":"Vistra Corp.",                  "role":"Nuclear and natural gas power for booming data-centre demand"},
        {"ticker":"ETR",   "company":"Entergy Corporation",           "role":"Nuclear utility serving the US South with growing load"},
        {"ticker":"EXC",   "company":"Exelon Corporation",            "role":"Largest regulated US utility with significant nuclear fleet"},
        {"ticker":"NRG",   "company":"NRG Energy Inc.",               "role":"Power retailer adding nuclear and gas capacity for AI load"},
        {"ticker":"AEP",   "company":"American Electric Power",       "role":"Transmission-heavy utility upgrading grid for industrial AI demand"},
        {"ticker":"SO",    "company":"Southern Company",              "role":"Owns Vogtle nuclear units 3 & 4, newest US reactors online"},
        {"ticker":"D",     "company":"Dominion Energy",               "role":"Virginia utility facing massive data-centre power demand"},
        {"ticker":"DUK",   "company":"Duke Energy Corporation",       "role":"Major Carolinas/Midwest utility investing in grid modernisation"},
        {"ticker":"GEV",   "company":"GE Vernova Inc.",               "role":"Gas turbines, wind and grid electrification equipment"},
        {"ticker":"PWR",   "company":"Quanta Services",               "role":"Electric grid construction and substation upgrades for AI era"},
        {"ticker":"MYR",   "company":"MYR Group Inc.",                "role":"Electrical contractor building transmission and distribution lines"},
        {"ticker":"ARRY",  "company":"Array Technologies",            "role":"Solar tracking systems enabling utility-scale renewable capacity"},
        {"ticker":"BE",    "company":"Bloom Energy Corporation",      "role":"Fuel-cell power systems for data-centre backup and baseload"},
        {"ticker":"OKLO",  "company":"Oklo Inc.",                     "role":"Advanced fission microreactor developer targeting data centres"},
        {"ticker":"SMR",   "company":"NuScale Power Corporation",     "role":"Designs small modular reactors (SMRs) for distributed power"},
        {"ticker":"X",     "company":"United States Steel",           "role":"Steel supply for nuclear plant and grid infrastructure builds"},
        {"ticker":"COP",   "company":"ConocoPhillips",                "role":"Natural gas supply bridging grid gap while nuclear comes online"},
        {"ticker":"LNG",   "company":"Cheniere Energy",               "role":"LNG export for global power-grid gas needs"},
        {"ticker":"BKR",   "company":"Baker Hughes Company",          "role":"Gas turbine services and small-modular-reactor component supply"},
        {"ticker":"HASI",  "company":"HA Sustainable Infrastructure", "role":"Finances clean-energy and grid modernisation projects"},
        {"ticker":"NEE",   "company":"NextEra Energy Inc.",           "role":"Largest US utility; wind, solar and storage at scale"},
        {"ticker":"AES",   "company":"AES Corporation",               "role":"Global utility accelerating clean-energy transition and storage"},
        {"ticker":"PCG",   "company":"PG&E Corporation",              "role":"California utility facing enormous EV and AI data-centre load"},
        {"ticker":"AWK",   "company":"American Water Works",          "role":"Critical water infrastructure adjacent to nuclear-plant cooling"},
    ],
    "Cybersecurity": [
        {"ticker":"CRWD",  "company":"CrowdStrike Holdings",          "role":"Falcon AI platform protecting endpoint, cloud and identity"},
        {"ticker":"PANW",  "company":"Palo Alto Networks",            "role":"Platformised network, cloud and SOC security suite"},
        {"ticker":"ZS",    "company":"Zscaler Inc.",                  "role":"Zero-trust cloud security proxy securing remote workforces"},
        {"ticker":"FTNT",  "company":"Fortinet Inc.",                 "role":"Firewall and SD-WAN security for mid-market and enterprise"},
        {"ticker":"S",     "company":"SentinelOne Inc.",              "role":"AI-native endpoint detection and response platform"},
        {"ticker":"CYBR",  "company":"CyberArk Software",            "role":"Privileged access management protecting critical credentials"},
        {"ticker":"OKTA",  "company":"Okta Inc.",                     "role":"Identity-as-a-service for workforce and customer authentication"},
        {"ticker":"QLYS",  "company":"Qualys Inc.",                   "role":"Cloud-based vulnerability management and compliance scanning"},
        {"ticker":"TENB",  "company":"Tenable Holdings",              "role":"Exposure management and vulnerability scanning across IT/OT"},
        {"ticker":"RPM",   "company":"RPM International",             "role":"Industrial coatings including protective products for critical infra"},
        {"ticker":"VRNS",  "company":"Varonis Systems",               "role":"Data security platform protecting unstructured enterprise data"},
        {"ticker":"ILMN",  "company":"Illumina Inc.",                 "role":"(genome data security adjacency) — see CHKP for direct fit"},
        {"ticker":"CHKP",  "company":"Check Point Software",          "role":"Network firewall and threat-prevention across cloud and on-premise"},
        {"ticker":"SAIL",  "company":"SailPoint Technologies",        "role":"AI-driven identity governance automating access across enterprises"},
        {"ticker":"ORCL",  "company":"Oracle Corporation",            "role":"Database security and cloud access management for enterprise"},
        {"ticker":"IBM",   "company":"IBM Corporation",               "role":"QRadar SIEM and managed security services for large enterprises"},
        {"ticker":"MSFT",  "company":"Microsoft Corporation",         "role":"Defender, Sentinel and Entra identity platform at scale"},
        {"ticker":"GOOGL", "company":"Alphabet Inc.",                 "role":"Google Mandiant threat intelligence and Chronicle SIEM"},
        {"ticker":"AMZN",  "company":"Amazon.com Inc.",               "role":"AWS security services — GuardDuty, Security Hub, IAM"},
        {"ticker":"NET",   "company":"Cloudflare Inc.",               "role":"Zero-trust network access and DDoS mitigation at global edge"},
        {"ticker":"GEN",   "company":"Gen Digital Inc.",              "role":"Consumer cybersecurity — Norton, Avast and LifeLock brands"},
        {"ticker":"NLOK",  "company":"NortonLifeLock",                "role":"Consumer identity protection and device security"},
        {"ticker":"LDOS",  "company":"Leidos Holdings",               "role":"Cybersecurity services for US government and defence agencies"},
        {"ticker":"BAH",   "company":"Booz Allen Hamilton",           "role":"Cyber consulting and managed detection for federal agencies"},
        {"ticker":"SAIC",  "company":"Science Applications Intl.",    "role":"IT and cyber services for US defence and intelligence community"},
    ],
    "Defense Technology & Autonomy": [
        {"ticker":"LMT",   "company":"Lockheed Martin Corporation",   "role":"F-35, missile systems and autonomous vehicle programmes"},
        {"ticker":"RTX",   "company":"RTX Corporation",               "role":"Patriot missiles, Raytheon sensors and Pratt & Whitney engines"},
        {"ticker":"NOC",   "company":"Northrop Grumman Corporation",  "role":"B-21 stealth bomber, space systems and autonomous drones"},
        {"ticker":"GD",    "company":"General Dynamics Corporation",  "role":"Gulfstream jets, Abrams tanks and Stryker armoured vehicles"},
        {"ticker":"BA",    "company":"Boeing Company",                "role":"F/A-18, P-8 maritime patrol and autonomous MQ-25 tanker"},
        {"ticker":"HII",   "company":"Huntington Ingalls Industries", "role":"Nuclear aircraft carriers and submarines for US Navy"},
        {"ticker":"LHX",   "company":"L3Harris Technologies",         "role":"ISR systems, electronic warfare and tactical radio networks"},
        {"ticker":"TDG",   "company":"TransDigm Group",               "role":"Proprietary aerospace components for military and commercial aircraft"},
        {"ticker":"LDOS",  "company":"Leidos Holdings",               "role":"AI-driven ISR analytics and autonomous vehicle integration"},
        {"ticker":"BAH",   "company":"Booz Allen Hamilton",           "role":"AI and data analytics for US defence and intelligence"},
        {"ticker":"KTOS",  "company":"Kratos Defense & Security",     "role":"Affordable tactical drones and hypersonic target systems"},
        {"ticker":"PLTR",  "company":"Palantir Technologies",         "role":"Maven Smart System AI for battlefield decision-making"},
        {"ticker":"RCAT",  "company":"Red Cat Holdings",              "role":"Small tactical drones for reconnaissance and strike missions"},
        {"ticker":"AVAV",  "company":"AeroVironment Inc.",            "role":"Switchblade loitering munitions and small UAS for US Army"},
        {"ticker":"SPCE",  "company":"Virgin Galactic",               "role":"Suborbital spaceflight platform with defence applications"},
        {"ticker":"RKLB",  "company":"Rocket Lab USA Inc.",           "role":"Small satellite launch and spacecraft for defence constellations"},
        {"ticker":"SPIR",  "company":"Spire Global Inc.",             "role":"Satellite data for weather, maritime and aviation intelligence"},
        {"ticker":"MAXR",  "company":"Maxar Technologies",            "role":"High-resolution satellite imagery for defence and intelligence"},
        {"ticker":"SAIC",  "company":"Science Applications Intl.",    "role":"IT modernisation and autonomous systems integration for DoD"},
        {"ticker":"CACI",  "company":"CACI International Inc.",       "role":"Intelligence analysis and cyber mission support for US agencies"},
        {"ticker":"DRS",   "company":"Leonardo DRS Inc.",             "role":"Electronic systems and sensors for US military platforms"},
        {"ticker":"ACHR",  "company":"Archer Aviation Inc.",          "role":"Electric VTOL aircraft for autonomous military logistics"},
        {"ticker":"JOBY",  "company":"Joby Aviation Inc.",            "role":"eVTOL air taxi with USAF agile logistics contract"},
        {"ticker":"ASTS",  "company":"AST SpaceMobile Inc.",          "role":"Direct-to-smartphone satellite network for resilient comms"},
        {"ticker":"MSTR",  "company":"MicroStrategy Inc.",            "role":"(replaced) — see PTC for defence digital twin fit"},
    ],
    "Robotics & Physical AI": [
        {"ticker":"ISRG",  "company":"Intuitive Surgical Inc.",       "role":"Da Vinci surgical robots dominating minimally-invasive surgery"},
        {"ticker":"ABB",   "company":"ABB Ltd. (US ADR)",             "role":"Industrial robots and cobots for factory automation globally"},
        {"ticker":"ROK",   "company":"Rockwell Automation",           "role":"Programmable logic controllers and factory automation software"},
        {"ticker":"EMR",   "company":"Emerson Electric Co.",          "role":"Process automation systems for industrial plants"},
        {"ticker":"HON",   "company":"Honeywell International",       "role":"Industrial automation, robotics and warehouse management systems"},
        {"ticker":"PTC",   "company":"PTC Inc.",                      "role":"Digital twin and IoT software connecting physical and AI systems"},
        {"ticker":"CGNX",  "company":"Cognex Corporation",            "role":"Machine vision systems guiding robot arms on factory lines"},
        {"ticker":"BRKS",  "company":"Brooks Automation",             "role":"Semiconductor wafer handling robots for chip fabrication"},
        {"ticker":"NVDA",  "company":"NVIDIA Corporation",            "role":"Isaac robotics platform and Jetson edge-AI for physical robots"},
        {"ticker":"GOOGL", "company":"Alphabet Inc.",                 "role":"DeepMind robotics research and Everyday Robots programme"},
        {"ticker":"AMZN",  "company":"Amazon.com Inc.",               "role":"Proteus, Sequoia and Digit warehouse robots at scale"},
        {"ticker":"TSLA",  "company":"Tesla Inc.",                    "role":"Optimus humanoid robot and Full Self-Driving AI stack"},
        {"ticker":"TER",   "company":"Teradyne Inc.",                 "role":"Collaborative robots (UR) and automated testing equipment"},
        {"ticker":"FLIR",  "company":"Teledyne FLIR (via TDY)",       "role":"Thermal imaging sensors enabling robot perception"},
        {"ticker":"TDY",   "company":"Teledyne Technologies",         "role":"Sensors and imaging systems for robotic and autonomous platforms"},
        {"ticker":"KEYS",  "company":"Keysight Technologies",         "role":"Test and measurement for autonomous vehicle and robot development"},
        {"ticker":"NXPI",  "company":"NXP Semiconductors",            "role":"Microcontrollers and radar chips for autonomous and robotic systems"},
        {"ticker":"STM",   "company":"STMicroelectronics",            "role":"Motor controllers and sensors powering industrial robots"},
        {"ticker":"FANUY", "company":"FANUC Corporation (US ADR)",    "role":"World's largest industrial robot maker for auto and electronics"},
        {"ticker":"IROBOT","company":"iRobot Corporation",            "role":"Consumer and commercial mobile robotics platforms"},
        {"ticker":"MBOT",  "company":"Microbot Medical Inc.",         "role":"Micro-robotics for minimally-invasive surgical procedures"},
        {"ticker":"ENPH",  "company":"Enphase Energy Inc.",           "role":"AI-driven microinverters and home energy automation systems"},
        {"ticker":"PATH",  "company":"UiPath Inc.",                   "role":"Software robots automating repetitive business processes at scale"},
        {"ticker":"AZTA",  "company":"Azenta Inc.",                   "role":"Automated sample storage and genomics robotics for life sciences"},
        {"ticker":"VICR",  "company":"Vicor Corporation",             "role":"Power modules enabling high-density power delivery in robots"},
    ],
    "GLP-1 / Metabolic Health": [
        {"ticker":"LLY",   "company":"Eli Lilly and Company",         "role":"Mounjaro and Zepbound — leading GLP-1 drugs for diabetes/obesity"},
        {"ticker":"NVO",   "company":"Novo Nordisk (US ADR)",         "role":"Ozempic and Wegovy — pioneer GLP-1 franchise with global scale"},
        {"ticker":"AZN",   "company":"AstraZeneca PLC (US ADR)",      "role":"GLP-1 pipeline and cardiovascular-metabolic drug portfolio"},
        {"ticker":"PFE",   "company":"Pfizer Inc.",                   "role":"Oral GLP-1 candidate danuglipron in late-stage trials"},
        {"ticker":"AMGN",  "company":"Amgen Inc.",                    "role":"MariTide monthly GLP-1 injection in Phase 3 development"},
        {"ticker":"RHHBY", "company":"Roche Holding (US ADR)",        "role":"CT-996 oral GLP-1/GIP agonist in early clinical stages"},
        {"ticker":"SNY",   "company":"Sanofi SA (US ADR)",            "role":"Metabolic disease drugs and GLP-1 partnership pipeline"},
        {"ticker":"VKTX",  "company":"Viking Therapeutics Inc.",       "role":"VK2735 oral GLP-1 showing strong Phase 2 weight-loss results"},
        {"ticker":"STVN",  "company":"Stevanato Group",               "role":"Drug delivery devices and containers for GLP-1 injectables"},
        {"ticker":"BDTX",  "company":"Blueprint Medicines",           "role":"Metabolic oncology pipeline adjacent to GLP-1 pathways"},
        {"ticker":"WST",   "company":"West Pharmaceutical Services",  "role":"Self-injection systems and containment for GLP-1 drug delivery"},
        {"ticker":"STE",   "company":"STERIS plc",                    "role":"Sterilisation and manufacturing services for GLP-1 drug makers"},
        {"ticker":"DXCM",  "company":"DexCom Inc.",                   "role":"Continuous glucose monitors essential for GLP-1 therapy management"},
        {"ticker":"ABT",   "company":"Abbott Laboratories",           "role":"FreeStyle Libre CGM and diagnostics for metabolic health monitoring"},
        {"ticker":"MDT",   "company":"Medtronic plc",                 "role":"Bariatric surgical devices and neuromodulation for obesity treatment"},
        {"ticker":"ISRG",  "company":"Intuitive Surgical Inc.",       "role":"Robotic bariatric surgery system for severe obesity cases"},
        {"ticker":"HUM",   "company":"Humana Inc.",                   "role":"Insurer with largest Medicare Advantage exposure to GLP-1 coverage"},
        {"ticker":"CI",    "company":"Cigna Group",                   "role":"Health insurer navigating GLP-1 formulary and coverage decisions"},
        {"ticker":"CVS",   "company":"CVS Health Corporation",        "role":"PBM and pharmacy dispensing and managing GLP-1 prescriptions"},
        {"ticker":"WBA",   "company":"Walgreens Boots Alliance",      "role":"Retail pharmacy and GLP-1 adherence programme provider"},
        {"ticker":"HIMS",  "company":"Hims & Hers Health Inc.",       "role":"Telehealth platform offering compounded semaglutide prescriptions"},
        {"ticker":"NRC",   "company":"National HealthCare Corp.",     "role":"Long-term care beneficiary as GLP-1 reduces obesity comorbidities"},
        {"ticker":"HOLX",  "company":"Hologic Inc.",                  "role":"Women's health diagnostics benefiting from metabolic health focus"},
        {"ticker":"TNDM",  "company":"Tandem Diabetes Care",          "role":"Insulin pumps used alongside GLP-1 in Type 1 diabetes management"},
        {"ticker":"ALGN",  "company":"Align Technology Inc.",         "role":"Dental clear aligners — oral health spending rises with GLP-1 users"},
    ],
}

SCREENER_THEMES = {k: k for k in THEME_TICKERS}


@app.route("/api/screener/themes")
def screener_themes():
    return jsonify(list(SCREENER_THEMES.keys()))


def _safe_round(val, ndigits=1):
    """Round a value, but return None instead of NaN/Inf (which break strict JSON.parse)."""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return round(f, ndigits)


@app.route("/api/screener/theme", methods=["POST"])
def screener_theme():
    """
    Use curated ticker lists per theme, then score each on Quality-at-a-Discount via yfinance.
    No API key required.
    """
    import pandas as pd
    body  = request.get_json(force=True) or {}
    theme = body.get("theme", "").strip()
    if theme not in THEME_TICKERS:
        return jsonify({"error": f"Unknown theme. Valid: {list(THEME_TICKERS.keys())}"}), 400

    candidates = THEME_TICKERS[theme]

    # ── yfinance metrics + scoring ────────────────────────────────
    results = []
    for item in candidates:
        ticker_sym = item.get("ticker", "").upper()
        company    = item.get("company", ticker_sym)
        role       = item.get("role", "—")
        if not ticker_sym:
            continue

        row = {
            "ticker": ticker_sym, "company": company, "role": role,
            "roic": None, "fcf_margin": None, "net_debt_ebitda": None,
            "rev_growth": None, "pct_off_52w": None,
            "trailing_pe": None, "forward_pe": None,
            "current_price": None, "market_cap": None, "sector": None,
        }

        try:
            t    = yf.Ticker(ticker_sym)
            info = t.info or {}

            row["current_price"] = info.get("currentPrice") or info.get("regularMarketPrice")
            row["market_cap"]    = info.get("marketCap")
            row["sector"]        = info.get("sector")
            row["trailing_pe"]   = info.get("trailingPE")
            row["forward_pe"]    = info.get("forwardPE")

            high_52w = info.get("fiftyTwoWeekHigh")
            cur_px   = row["current_price"]
            if high_52w and cur_px and high_52w > 0 and not math.isnan(high_52w):
                row["pct_off_52w"] = _safe_round((high_52w - cur_px) / high_52w * 100, 1)

            try:
                fin = t.financials
                rev_key = next((k for k in ['Total Revenue','Revenue'] if k in fin.index), None)
                if rev_key and len(fin.columns) >= 2:
                    r0, r1 = fin.loc[rev_key].iloc[0], fin.loc[rev_key].iloc[1]
                    if r1 and r1 != 0 and not (pd.isna(r0) or pd.isna(r1)):
                        row["rev_growth"] = _safe_round((r0 - r1) / abs(r1) * 100, 1)
            except Exception:
                pass

            try:
                fin = t.financials
                bs  = t.balance_sheet
                ebit_key = next((k for k in ['EBIT','Operating Income','Total Operating Income As Reported'] if k in fin.index), None)
                if ebit_key:
                    ebit   = fin.loc[ebit_key].iloc[0]
                    ta_key = next((k for k in ['Total Assets'] if k in bs.index), None)
                    cl_key = next((k for k in ['Current Liabilities','Total Current Liabilities'] if k in bs.index), None)
                    if ta_key and cl_key and not pd.isna(ebit):
                        inv_cap = bs.loc[ta_key].iloc[0] - bs.loc[cl_key].iloc[0]
                        if inv_cap and inv_cap > 0 and not pd.isna(inv_cap):
                            row["roic"] = _safe_round(ebit * 0.79 / inv_cap * 100, 1)
            except Exception:
                pass

            try:
                cf   = t.cashflow
                fin2 = t.financials
                ocf_key   = next((k for k in ['Operating Cash Flow','Total Cash From Operating Activities'] if k in cf.index), None)
                capex_key = next((k for k in ['Capital Expenditure','Capital Expenditures'] if k in cf.index), None)
                rev_key2  = next((k for k in ['Total Revenue','Revenue'] if k in fin2.index), None)
                if ocf_key:
                    ocf   = cf.loc[ocf_key].iloc[0]
                    capex = cf.loc[capex_key].iloc[0] if capex_key else 0
                    fcf   = ocf + capex
                    row["_fcf_positive"] = bool(fcf > 0) if not pd.isna(fcf) else None
                    if rev_key2:
                        rev = fin2.loc[rev_key2].iloc[0]
                        if rev and rev != 0 and not (pd.isna(fcf) or pd.isna(rev)):
                            row["fcf_margin"] = _safe_round(fcf / rev * 100, 1)
            except Exception:
                row["_fcf_positive"] = None

            try:
                ebitda   = info.get("ebitda")
                net_debt = (info.get("totalDebt") or 0) - (info.get("totalCash") or 0)
                if ebitda and ebitda > 0 and not math.isnan(ebitda):
                    row["net_debt_ebitda"] = _safe_round(net_debt / ebitda, 2)
            except Exception:
                pass

        except Exception:
            pass

        # Final safety net: scrub any stray NaN/Inf floats before serializing
        for k, v in list(row.items()):
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                row[k] = None

        # ── Scoring ──────────────────────────────────────────────
        score = 0
        if row["roic"] is not None and row["roic"] >= 15:          score += 20
        if row.get("_fcf_positive") is True:                        score += 20
        if row["net_debt_ebitda"] is not None and row["net_debt_ebitda"] < 2: score += 20
        if row["rev_growth"] is not None and row["rev_growth"] > 0: score += 20
        fpe, tpe = row["forward_pe"], row["trailing_pe"]
        if row["pct_off_52w"] is not None and row["pct_off_52w"] >= 15: score += 10
        if fpe and tpe and fpe > 0 and tpe > 0 and fpe < tpe:       score += 10

        row["score"] = score
        row["tier"]  = "BEST" if score >= 80 else "STRONG" if score >= 65 else "WATCH" if score >= 50 else "AVOID"
        row.pop("_fcf_positive", None)
        results.append(row)

    results.sort(key=lambda x: x["score"], reverse=True)
    for i, r in enumerate(results, 1):
        r["rank"] = i

    return jsonify({"theme": theme, "tickers": results})


# ═══════════════════════════════════════════════
# RECOMMENDATIONS
# ═══════════════════════════════════════════════

@app.route("/api/recommend/<ticker>")
def recommendation(ticker):
    risk   = request.args.get("risk", "medium")
    result = recommend(ticker.upper(), risk_tolerance=risk)
    return jsonify(result)


# ═══════════════════════════════════════════════
# ANALYSIS
# ═══════════════════════════════════════════════

@app.route("/api/analyse/<ticker>/<strategy>")
def analysis(ticker, strategy):
    result = analyse(ticker.upper(), strategy)
    return jsonify(result)


# ═══════════════════════════════════════════════
# BACKTEST
# ═══════════════════════════════════════════════

@app.route("/api/backtest", methods=["POST"])
def backtest():
    body     = request.get_json() or {}
    ticker   = body.get("ticker", "").upper()
    strategy = body.get("strategy", "covered_call")
    period   = body.get("period", "2y")
    dte      = int(body.get("dte", 30))
    otm_pct  = float(body.get("otm_pct", 0.05))
    width_mode   = body.get("width_mode", "pct")
    strike_width = float(body.get("strike_width", 5.0))
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    result = run_backtest(ticker, strategy, period=period, dte=dte, otm_pct=otm_pct,
                           width_mode=width_mode, strike_width=strike_width)
    if "error" not in result:
        run_id = save_backtest(ticker, strategy, body, result["summary"])
        result["run_id"] = run_id
    return jsonify(result)


@app.route("/api/backtest/history")
def backtest_history():
    return jsonify(list_backtests(request.args.get("ticker")))


@app.route("/api/backtest/<int:run_id>")
def backtest_detail(run_id):
    bt = get_backtest(run_id)
    if bt is None:
        abort(404)
    return jsonify(bt)


# ═══════════════════════════════════════════════
# PLATFORMS
# ═══════════════════════════════════════════════

@app.route("/api/platforms")
def platforms_list():
    category = request.args.get("category")
    return jsonify(list_platforms(category))


@app.route("/api/platforms", methods=["POST"])
def platform_add():
    body = request.get_json() or {}
    name = body.get("name", "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    new_id = add_platform(name, body.get("category", "all"))
    return jsonify({"id": new_id, "name": name}), 201


@app.route("/api/platforms/<int:platform_id>", methods=["DELETE"])
def platform_delete(platform_id):
    delete_platform(platform_id)
    return jsonify({"status": "deleted"})


# ═══════════════════════════════════════════════
# EQUITIES
# ═══════════════════════════════════════════════

EQUITY_IMPORT_FIELDS  = ["ticker", "company", "shares", "cost_price", "purchase_date",
                          "sector", "industry", "platform", "currency", "notes"]
EQUITY_REPORT_FIELDS  = ["id", "ticker", "company", "sector", "industry", "platform",
                          "shares", "cost_price", "current_price", "price_date",
                          "market_value_usd", "market_value_sgd",
                          "unrealized_pnl_usd", "unrealized_pnl_sgd", "unrealized_pnl_pct",
                          "purchase_date", "currency", "notes"]


@app.route("/api/equities", methods=["GET"])
def equities_list():
    rows = list_equities()
    # Batch-fetch all unique FX pairs in one network call
    ccys = {(r.get("currency") or "USD").upper() for r in rows}
    pairs = ["USDSGD"] + [f"{c}SGD" for c in ccys if c not in ("USD", "SGD")]
    _batch_fx_rates(list(dict.fromkeys(pairs)))
    fx       = get_fx_rate()
    holdings = [_enrich_equity(e, fx) for e in rows]
    return jsonify({"holdings": holdings, "fx_usdsgd": fx})


@app.route("/api/equities", methods=["POST"])
def equity_add():
    data   = request.get_json() or {}
    ticker = data.get("ticker", "").strip().upper()
    if not ticker:
        return jsonify({"error": "ticker is required"}), 400
    data["ticker"] = ticker
    # Auto-fetch company info + latest price
    try:
        t    = yf.Ticker(ticker)
        info = _safe_yf_info(ticker)
        data.setdefault("company",  info.get("longName"))
        data.setdefault("sector",   info.get("sector"))
        data.setdefault("industry", info.get("industry"))
        if not data.get("current_price"):
            hist = t.history(period="1d")
            if not hist.empty:
                data["current_price"] = round(float(hist["Close"].iloc[-1]), 4)
                data["price_date"]    = datetime.today().strftime("%Y-%m-%d")
    except Exception:
        pass
    new_id = add_equity(data)
    return jsonify({"id": new_id, "status": "created"}), 201


@app.route("/api/equities/<int:eq_id>", methods=["GET"])
def equity_get(eq_id):
    eq = get_equity(eq_id)
    if eq is None:
        abort(404)
    return jsonify(_enrich_equity(eq))


@app.route("/api/equities/<int:eq_id>", methods=["PUT"])
def equity_update(eq_id):
    update_equity(eq_id, request.get_json() or {})
    return jsonify({"status": "updated", "equity": _enrich_equity(get_equity(eq_id))})


@app.route("/api/equities/<int:eq_id>", methods=["DELETE"])
def equity_delete(eq_id):
    delete_equity(eq_id)
    return jsonify({"status": "deleted"})


@app.route("/api/equities/bulk-delete", methods=["POST"])
def equity_bulk_delete():
    data = request.get_json(force=True)
    ids  = data.get("ids", [])
    if ids == "all":
        rows = list_equities()
        ids  = [r["id"] for r in rows]
    deleted = 0
    for eid in ids:
        delete_equity(int(eid))
        deleted += 1
    return jsonify({"deleted": deleted})


def _get_live_price(ticker):
    """Most current price available for a ticker.
    Prefers the live quote (fast_info.last_price), which Yahoo updates
    continuously during market hours. Falls back to the daily history
    endpoint, which can lag a day or more in posting the latest session's
    close. Returns (price, date_str, source) or (None, None, None)."""
    try:
        fi = yf.Ticker(ticker).fast_info
        price = fi.get("last_price") or fi.get("lastPrice")
        if price:
            return round(float(price), 4), datetime.now().strftime("%Y-%m-%d"), "quote"
    except Exception:
        pass
    try:
        hist  = yf.Ticker(ticker).history(period="5d")
        close = hist["Close"].dropna() if not hist.empty else None
        if close is not None and not close.empty:
            return round(float(close.iloc[-1]), 4), close.index[-1].strftime("%Y-%m-%d"), "history"
    except Exception:
        pass
    return None, None, None


@app.route("/api/equities/<int:eq_id>/refresh-price", methods=["POST"])
def equity_refresh_price(eq_id):
    eq = get_equity(eq_id)
    if eq is None:
        abort(404)
    price, date, source = _get_live_price(eq["ticker"])
    if price is None:
        return jsonify({"error": "No price data returned"}), 502
    update_equity_price(eq_id, price, date)
    return jsonify({"ticker": eq["ticker"], "price": price, "date": date, "source": source})


@app.route("/api/equities/refresh-all", methods=["POST"])
def equities_refresh_all():
    results = []
    for eq in list_equities():
        try:
            price, date, source = _get_live_price(eq["ticker"])
            if price is not None:
                update_equity_price(eq["id"], price, date)
                results.append({"ticker": eq["ticker"], "price": price, "date": date, "source": source, "ok": True})
            else:
                results.append({"ticker": eq["ticker"], "ok": False, "error": "No data"})
        except Exception as e:
            results.append({"ticker": eq["ticker"], "ok": False, "error": str(e)})
    return jsonify(results)


@app.route("/api/equities/<int:eq_id>/set-price", methods=["POST"])
def equity_set_price(eq_id):
    data  = request.get_json() or {}
    price = data.get("price")
    if price is None:
        return jsonify({"error": "price required"}), 400
    date = data.get("date", datetime.today().strftime("%Y-%m-%d"))
    update_equity_price(eq_id, float(price), date)
    return jsonify({"status": "updated", "price": float(price), "date": date})


# ── Purchase history ──────────────────────────

@app.route("/api/equities/<int:eq_id>/purchases", methods=["GET"])
def purchases_list(eq_id):
    return jsonify(list_purchases(eq_id))


@app.route("/api/equities/<int:eq_id>/purchases", methods=["POST"])
def purchase_add(eq_id):
    """Add a new lot. Auto-recalculates weighted avg cost on parent equity."""
    data   = request.get_json() or {}
    shares = data.get("shares")
    price  = data.get("price")
    if not shares or not price:
        return jsonify({"error": "shares and price required"}), 400
    updated = add_purchase(
        eq_id, float(shares), float(price),
        date=data.get("date"), platform=data.get("platform"), notes=data.get("notes")
    )
    return jsonify({"status": "added", "equity": _enrich_equity(updated)}), 201


@app.route("/api/purchases/<int:purchase_id>", methods=["DELETE"])
def purchase_delete(purchase_id):
    delete_purchase(purchase_id)
    return jsonify({"status": "deleted"})


# ── CSV / XLSX import + export ────────────────

@app.route("/api/equities/template", methods=["GET"])
def equities_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EQUITY_IMPORT_FIELDS)
    writer.writeheader()
    # One ticker, two platforms — each row becomes its own DB entry and they appear grouped in the UI
    writer.writerow({
        "ticker": "AAPL", "company": "Apple Inc.", "shares": "50",
        "cost_price": "150.00", "purchase_date": "2024-01-15",
        "sector": "Technology", "industry": "Consumer Electronics",
        "platform": "IB", "currency": "USD", "notes": "AAPL on IB"
    })
    writer.writerow({
        "ticker": "AAPL", "company": "Apple Inc.", "shares": "30",
        "cost_price": "170.00", "purchase_date": "2024-06-01",
        "sector": "Technology", "industry": "Consumer Electronics",
        "platform": "Webull", "currency": "USD", "notes": "Same ticker different platform - will be grouped in UI"
    })
    writer.writerow({
        "ticker": "0700.HK", "company": "Tencent Holdings", "shares": "200",
        "cost_price": "380.00", "purchase_date": "2024-03-01",
        "sector": "Communication Services", "industry": "Internet Content",
        "platform": "IB", "currency": "HKD", "notes": "HK stock example"
    })
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=equities_import_template.csv"})


@app.route("/api/equities/import", methods=["POST"])
def equities_import():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    try:
        raw_rows = _read_uploaded_file(f)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    rows, errors = [], []
    for i, row in enumerate(raw_rows, start=2):
        ticker = str(row.get("ticker") or "").strip().upper()
        if not ticker:
            errors.append(f"Row {i}: ticker blank"); continue
        try:
            parsed = {
                "ticker"       : ticker,
                "company"      : str(row.get("company") or "").strip() or None,
                "shares"       : float(row.get("shares") or 0),
                "cost_price"   : float(row.get("cost_price") or 0),
                "purchase_date": str(row.get("purchase_date") or "").strip() or None,
                "platform"     : str(row.get("platform") or "").strip() or None,
                "sector"       : str(row.get("sector") or "").strip() or None,
                "industry"     : str(row.get("industry") or "").strip() or None,
                "currency"     : str(row.get("currency") or "USD").strip() or "USD",
                "notes"        : str(row.get("notes") or "").strip() or None,
            }
            # Auto-fetch company info (+ sector/industry if blank)
            try:
                info = _safe_yf_info(ticker)
                if not parsed["company"]:
                    parsed["company"] = info.get("longName")
                if not parsed["sector"]:
                    parsed["sector"]   = info.get("sector")
                if not parsed["industry"]:
                    parsed["industry"] = info.get("industry")
                # Also grab latest price
                hist = yf.Ticker(ticker).history(period="1d")
                if not hist.empty:
                    parsed["current_price"] = round(float(hist["Close"].iloc[-1]), 4)
                    parsed["price_date"]    = datetime.today().strftime("%Y-%m-%d")
            except Exception:
                pass
            rows.append(parsed)
        except ValueError as e:
            errors.append(f"Row {i} ({ticker}): {e}")

    count, bulk_errors = bulk_insert_equities(rows)
    errors.extend(bulk_errors)
    return jsonify({"imported": count, "errors": errors})


@app.route("/api/equities/refresh-sectors", methods=["POST"])
def refresh_equity_sectors():
    """Fetch and save sector/industry for any equity that has no sector stored."""
    equities = list_equities()
    updated, failed = 0, 0
    for eq in equities:
        if eq.get("sector"):          # already has one — skip
            continue
        try:
            info = _fetch_info(eq["ticker"])
            if info.get("sector"):
                patch = {"sector": info["sector"]}
                if not eq.get("company") and info.get("company"):
                    patch["company"] = info["company"]
                if not eq.get("industry") and info.get("industry"):
                    patch["industry"] = info["industry"]
                update_equity(eq["id"], patch)
                updated += 1
        except Exception:
            failed += 1
    return jsonify({"updated": updated, "failed": failed})


@app.route("/api/equities/report", methods=["GET"])
def equities_report():
    fx       = get_fx_rate()
    holdings = [_enrich_equity(e, fx) for e in list_equities()]
    output   = io.StringIO()
    writer   = csv.DictWriter(output, fieldnames=EQUITY_REPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(holdings)
    summary = equities_summary()
    writer.writerow({
        "ticker"            : "TOTAL",
        "unrealized_pnl_sgd": summary["total_pnl"],
        "unrealized_pnl_pct": summary["total_pnl_pct"],
        "market_value_sgd"  : round(summary["total_value"] * fx, 2),
        "market_value_usd"  : summary["total_value"],
    })
    output.seek(0)
    filename = f"equities_report_{datetime.today().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ═══════════════════════════════════════════════
# CRYPTO
# ═══════════════════════════════════════════════

CRYPTO_IMPORT_FIELDS = ["symbol", "quantity", "cost_price", "purchase_date",
                         "platform", "currency", "notes"]
CRYPTO_REPORT_FIELDS = ["id", "symbol", "name", "quantity", "cost_price",
                         "current_price", "price_date", "platform",
                         "market_value_usd", "market_value_sgd",
                         "unrealized_pnl_usd", "unrealized_pnl_sgd",
                         "unrealized_pnl_pct", "purchase_date", "currency", "notes"]


@app.route("/api/crypto", methods=["GET"])
def crypto_list():
    fx       = get_fx_rate()
    holdings = [_enrich_crypto(c, fx) for c in list_crypto()]
    return jsonify({"holdings": holdings, "fx_usdsgd": fx})


@app.route("/api/crypto", methods=["POST"])
def crypto_add():
    data   = request.get_json() or {}
    symbol = data.get("symbol", "").strip().upper()
    if not symbol:
        return jsonify({"error": "symbol is required"}), 400
    data["symbol"] = symbol
    # Auto-fetch current price from yfinance (symbol-USD format)
    if not data.get("current_price"):
        try:
            hist = yf.Ticker(f"{symbol}-USD").history(period="1d")
            if not hist.empty:
                data["current_price"] = round(float(hist["Close"].iloc[-1]), 4)
                data["price_date"]    = datetime.today().strftime("%Y-%m-%d")
            # Fetch name
            info = _safe_yf_info(f"{symbol}-USD")
            data.setdefault("name", info.get("longName") or info.get("shortName"))
        except Exception:
            pass
    new_id = add_crypto(data)
    return jsonify({"id": new_id, "status": "created"}), 201


@app.route("/api/crypto/<int:cr_id>", methods=["GET"])
def crypto_get(cr_id):
    cr = get_crypto(cr_id)
    if cr is None:
        abort(404)
    return jsonify(_enrich_crypto(cr))


@app.route("/api/crypto/<int:cr_id>", methods=["PUT"])
def crypto_update(cr_id):
    update_crypto(cr_id, request.get_json() or {})
    return jsonify({"status": "updated"})


@app.route("/api/crypto/<int:cr_id>", methods=["DELETE"])
def crypto_delete(cr_id):
    delete_crypto(cr_id)
    return jsonify({"status": "deleted"})


@app.route("/api/crypto/bulk-delete", methods=["POST"])
def crypto_bulk_delete():
    data = request.get_json(force=True)
    ids  = data.get("ids", [])
    if ids == "all":
        rows = list_crypto()
        ids  = [r["id"] for r in rows]
    deleted = 0
    for cid in ids:
        delete_crypto(int(cid))
        deleted += 1
    return jsonify({"deleted": deleted})


@app.route("/api/crypto/<int:cr_id>/refresh-price", methods=["POST"])
def crypto_refresh_price(cr_id):
    cr = get_crypto(cr_id)
    if cr is None:
        abort(404)
    try:
        hist  = yf.Ticker(f"{cr['symbol']}-USD").history(period="2d")
        if hist.empty:
            return jsonify({"error": "No price data"}), 502
        price = round(float(hist["Close"].iloc[-1]), 4)
        date  = hist.index[-1].strftime("%Y-%m-%d")
        update_crypto_price(cr_id, price, date)
        return jsonify({"symbol": cr["symbol"], "price": price, "date": date})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/crypto/refresh-all", methods=["POST"])
def crypto_refresh_all():
    results = []
    for cr in list_crypto():
        try:
            hist  = yf.Ticker(f"{cr['symbol']}-USD").history(period="2d")
            if not hist.empty:
                price = round(float(hist["Close"].iloc[-1]), 4)
                date  = hist.index[-1].strftime("%Y-%m-%d")
                update_crypto_price(cr["id"], price, date)
                results.append({"symbol": cr["symbol"], "price": price, "ok": True})
            else:
                results.append({"symbol": cr["symbol"], "ok": False, "error": "No data"})
        except Exception as e:
            results.append({"symbol": cr["symbol"], "ok": False, "error": str(e)})
    return jsonify(results)


@app.route("/api/crypto/template", methods=["GET"])
def crypto_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CRYPTO_IMPORT_FIELDS)
    writer.writeheader()
    writer.writerow({"symbol": "BTC", "quantity": "0.5", "cost_price": "40000",
                     "purchase_date": "2024-01-15", "platform": "Binance",
                     "currency": "USD", "notes": "Example"})
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=crypto_import_template.csv"})


@app.route("/api/crypto/import", methods=["POST"])
def crypto_import():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        raw_rows = _read_uploaded_file(request.files["file"])
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    count, errors = 0, []
    for i, row in enumerate(raw_rows, start=2):
        symbol = str(row.get("symbol") or "").strip().upper()
        if not symbol:
            errors.append(f"Row {i}: symbol blank"); continue
        try:
            parsed = {
                "symbol"       : symbol,
                "quantity"     : float(row.get("quantity") or 0),
                "cost_price"   : float(row.get("cost_price") or 0),
                "purchase_date": str(row.get("purchase_date") or "").strip() or None,
                "platform"     : str(row.get("platform") or "").strip() or None,
                "currency"     : str(row.get("currency") or "USD").strip() or "USD",
                "notes"        : str(row.get("notes") or "").strip() or None,
            }
            # Auto-fetch price + name
            try:
                hist = yf.Ticker(f"{symbol}-USD").history(period="1d")
                if not hist.empty:
                    parsed["current_price"] = round(float(hist["Close"].iloc[-1]), 4)
                    parsed["price_date"]    = datetime.today().strftime("%Y-%m-%d")
                info = _safe_yf_info(f"{symbol}-USD")
                parsed["name"] = info.get("longName") or info.get("shortName")
            except Exception:
                pass
            add_crypto(parsed)
            count += 1
        except Exception as e:
            errors.append(f"Row {i} ({symbol}): {e}")
    return jsonify({"imported": count, "errors": errors})


@app.route("/api/crypto/report", methods=["GET"])
def crypto_report():
    fx       = get_fx_rate()
    holdings = [_enrich_crypto(c, fx) for c in list_crypto()]
    output   = io.StringIO()
    writer   = csv.DictWriter(output, fieldnames=CRYPTO_REPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(holdings)
    output.seek(0)
    filename = f"crypto_report_{datetime.today().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ═══════════════════════════════════════════════
# CASH
# ═══════════════════════════════════════════════

def _enrich_cash(row: dict, usdsgd: float) -> dict:
    """Add amount_sgd field using live FX rates."""
    ccy    = (row.get("currency") or "SGD").upper()
    amount = row.get("amount") or 0
    rate   = _to_sgd_rate(ccy, usdsgd)
    return {**row, "amount_sgd": round(amount * rate, 2), "fx_rate": rate}


@app.route("/api/cash", methods=["GET"])
def cash_list():
    rows    = list_cash()
    usdsgd  = get_fx_rate("USD", "SGD")
    ccys    = {(r.get("currency") or "SGD").upper() for r in rows if r.get("currency", "SGD") != "SGD"}
    if ccys:
        _batch_fx_rates([f"{c}SGD" for c in ccys if c != "USD"] + (["USDSGD"] if "USD" in ccys else []))
    enriched = [_enrich_cash(r, usdsgd) for r in rows]
    return jsonify({"cash": enriched, "fx_usdsgd": usdsgd})


@app.route("/api/cash", methods=["POST"])
def cash_add():
    data = request.get_json() or {}
    if not data.get("platform") or data.get("amount") is None:
        return jsonify({"error": "platform and amount required"}), 400
    new_id = add_cash(data)
    return jsonify({"id": new_id, "status": "created"}), 201


@app.route("/api/cash/<int:cash_id>", methods=["PUT"])
def cash_update(cash_id):
    update_cash(cash_id, request.get_json() or {})
    return jsonify({"status": "updated"})


@app.route("/api/cash/<int:cash_id>", methods=["DELETE"])
def cash_delete(cash_id):
    delete_cash(cash_id)
    return jsonify({"status": "deleted"})


# ═══════════════════════════════════════════════
# OPTIONS TRADES
# ═══════════════════════════════════════════════

OPTIONS_IMPORT_FIELDS = ["ticker", "strategy", "direction", "option_type",
                          "strike", "expiry", "contracts", "premium", "open_date", "platform", "notes"]
OPTIONS_REPORT_FIELDS = ["id", "ticker", "strategy", "direction", "option_type",
                          "strike", "expiry", "contracts", "premium", "open_date",
                          "close_date", "close_premium", "status", "pnl", "notes"]


@app.route("/api/trades", methods=["GET"])
def trades_list():
    return jsonify(list_trades(request.args.get("status")))


@app.route("/api/trades", methods=["POST"])
def trade_add():
    data   = request.get_json() or {}
    new_id = add_trade(data)
    return jsonify({"id": new_id, "status": "created"}), 201


@app.route("/api/trades/<int:trade_id>", methods=["GET"])
def trade_get(trade_id):
    trade = get_trade(trade_id)
    if trade is None:
        abort(404)
    return jsonify(trade)


@app.route("/api/trades/<int:trade_id>", methods=["PUT"])
def trade_update(trade_id):
    data = request.get_json() or {}
    update_trade(trade_id, data)
    return jsonify({"status": "updated", "trade": get_trade(trade_id)})


@app.route("/api/trades/<int:trade_id>/close", methods=["POST"])
def trade_close(trade_id):
    data          = request.get_json() or {}
    close_premium = data.get("close_premium")
    if close_premium is None:
        return jsonify({"error": "close_premium required"}), 400
    status = data.get("status", "CLOSED")  # supports EXPIRED
    close_trade(trade_id, float(close_premium), data.get("close_date"), status)
    return jsonify({"status": "closed", "trade": get_trade(trade_id)})


@app.route("/api/trades/<int:trade_id>", methods=["DELETE"])
def trade_delete(trade_id):
    delete_trade(trade_id)
    return jsonify({"status": "deleted"})


@app.route("/api/trades/bulk-delete", methods=["POST"])
def trades_bulk_delete():
    data = request.get_json(force=True)
    ids  = data.get("ids", [])
    if ids == "all":
        rows = list_trades()
        ids  = [r["id"] for r in rows]
    deleted = 0
    for tid in ids:
        delete_trade(int(tid))
        deleted += 1
    return jsonify({"deleted": deleted})


@app.route("/api/trades/template", methods=["GET"])
def trades_template():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=OPTIONS_IMPORT_FIELDS)
    writer.writeheader()
    writer.writerow({"ticker": "AAPL", "strategy": "Covered Call", "direction": "SELL",
                     "option_type": "CALL", "strike": "200", "expiry": "2025-08-15",
                     "contracts": "1", "premium": "3.50", "open_date": "2025-07-18",
                     "platform": "IB", "notes": "Example"})
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=options_import_template.csv"})


@app.route("/api/trades/import", methods=["POST"])
def trades_import():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        raw_rows = _read_uploaded_file(request.files["file"])
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    imported, errors = 0, []
    for i, row in enumerate(raw_rows, start=2):
        ticker = (row.get("ticker") or "").strip().upper()
        if not ticker:
            errors.append(f"Row {i}: ticker blank"); continue
        try:
            add_trade({
                "ticker"     : ticker,
                "strategy"   : (row.get("strategy") or "").strip(),
                "direction"  : (row.get("direction") or "SELL").strip().upper(),
                "option_type": (row.get("option_type") or "CALL").strip().upper(),
                "strike"     : float(row.get("strike") or 0),
                "expiry"     : (row.get("expiry") or "").strip(),
                "contracts"  : int(float(row.get("contracts") or 1)),
                "premium"    : float(row.get("premium") or 0),
                "open_date"  : (row.get("open_date") or "").strip(),
                "platform"   : (row.get("platform") or "").strip() or None,
                "notes"      : (row.get("notes") or "").strip() or None,
            })
            imported += 1
        except Exception as e:
            errors.append(f"Row {i} ({ticker}): {e}")
    return jsonify({"imported": imported, "skipped": len(errors), "errors": errors})


@app.route("/api/trades/report", methods=["GET"])
def trades_report():
    trades = list_trades()
    output = io.StringIO()
    fields = ["id","ticker","strategy","direction","option_type","strike","expiry",
              "contracts","premium","open_date","close_date","close_premium",
              "realized_pnl","status","notes"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader(); writer.writerows(trades)
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=options_report.csv"})


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@app.route("/api/dashboard")
def dashboard():
    fx = get_fx_rate("USD", "SGD")
    equities = list_equities()
    enriched = [_enrich_equity(e, fx) for e in equities]
    eq_total_val  = sum(e["market_value_sgd"]   for e in enriched)   # SGD
    eq_total_cost = sum(e["cost_basis_sgd"]     for e in enriched)   # SGD
    eq_total_pnl  = sum(e["unrealized_pnl_sgd"] for e in enriched)   # SGD

    crypto_rows = list_crypto()
    cr_total_val = sum((c.get("current_price") or 0) * (c.get("quantity") or 0) for c in crypto_rows)

    trades = list_trades()

    # Group legs into strategies.
    # Priority: explicit trade_group → auto-group by (ticker, strategy, expiry, open_date)
    # so that multi-leg spreads entered without a trade_group (e.g. two-leg Bull Put Spread)
    # are still collapsed into one strategy for KPI purposes.
    def _auto_key(t):
        exp      = (t.get("expiry")     or "").split(" ")[0]
        open_dt  = (t.get("open_date")  or "").split(" ")[0]
        return f"{t['ticker']}|{t.get('strategy') or ''}|{exp}|{open_dt}"

    _groups: dict = {}
    for t in trades:
        key = t.get("trade_group") or _auto_key(t)
        _groups.setdefault(key, []).append(t)

    DONE = {"CLOSED", "EXPIRED", "ASSIGNED", "EXERCISED"}

    def _group_status(legs):
        statuses = {(l.get("status") or "OPEN").upper() for l in legs}
        if any(s not in DONE for s in statuses):
            return "OPEN"
        return "CLOSED" if statuses <= DONE else "PARTIAL"

    def _group_pnl(legs):
        return sum(float(l.get("pnl") or 0) for l in legs)

    def _group_close_date(legs):
        dates = [l.get("close_date") or l.get("expiry") or "" for l in legs]
        return max((d for d in dates if d), default="")

    open_strategies   = [g for g in _groups.values() if _group_status(g) == "OPEN"]
    closed_strategies = [g for g in _groups.values() if _group_status(g) == "CLOSED"]

    total_pnl = sum(_group_pnl(g) for g in _groups.values())  # all realized P&L incl. partial
    wins      = sum(1 for g in closed_strategies if _group_pnl(g) > 0)
    win_rate  = round(wins / len(closed_strategies) * 100, 1) if closed_strategies else 0

    # Keep open_t for the open-trades table (show individual legs for detail)
    open_t = [leg for g in open_strategies for leg in g]

    watchlist = list_watchlist()
    recent_bt = get_recent_backtests(5) if callable(globals().get("get_recent_backtests")) else []

    return jsonify({
        "fx_usdsgd": fx,
        "portfolio": {
            "open_positions": len(open_strategies),   # strategy count, not leg count
            "total_trades":   len(_groups),            # strategy count
            "total_pnl":      round(total_pnl, 2),
            "win_rate":       win_rate,
            "equities": {
                "total_holdings": len(equities),
                "total_value":    round(eq_total_val, 2),    # SGD
                "total_cost":     round(eq_total_cost, 2),   # SGD
                "total_pnl_sgd":  round(eq_total_pnl, 2),   # SGD
            },
            "crypto": {
                "total_holdings": len(crypto_rows),
                "total_value":    round(cr_total_val, 2),
            },
        },
        "open_trades":       open_t[:10],
        "watchlist":         watchlist[:10],
        "recent_backtests":  recent_bt,
    })


# ─── PERFORMANCE CHART ────────────────────────────────────────────────────────

_perf_cache: dict = {}

@app.route("/api/dashboard/performance")
def dashboard_performance():
    import pandas as pd
    period    = request.args.get("period", "6mo")
    cache_key = f"perf_{period}"
    cached    = _perf_cache.get(cache_key)
    if cached and (datetime.now() - cached["ts"]).seconds < 300:
        return jsonify(cached["data"])
    fx       = get_fx_rate("USD", "SGD")
    holdings = list_equities()
    if not holdings:
        return jsonify({"dates": [], "portfolio_sgd": [], "cost_sgd": [], "sectors": {}})
    ticker_info = {}
    for h in holdings:
        lots = list_purchases(h["id"])
        if not lots:
            lots = [{"date": (h.get("purchase_date") or "2020-01-01")[:10],
                     "shares": h.get("shares", 0), "price": h.get("cost_price", 0)}]
        ticker_info[h["ticker"]] = {
            "lots":   sorted(lots, key=lambda l: (l.get("date") or "2000-01-01")[:10]),
            "to_sgd": _to_sgd_rate((h.get("currency") or "USD").upper(), fx),
        }
    tickers = list(ticker_info.keys())
    try:
        raw = yf.download(tickers, period=period, auto_adjust=True, progress=False, group_by="ticker")
        if raw.empty:
            return jsonify({"dates": [], "portfolio_sgd": [], "cost_sgd": [], "sectors": {}})
        if len(tickers) == 1:
            price_df = pd.DataFrame({tickers[0]: raw["Close"]})
        else:
            closes = {}
            for t in tickers:
                try: closes[t] = raw[t]["Close"]
                except (KeyError, TypeError): pass
            price_df = pd.DataFrame(closes)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    price_df = price_df.resample("W").last().ffill()
    dates_out, portfolio_out, cost_out = [], [], []
    for idx in price_df.index:
        date_str = idx.strftime("%Y-%m-%d")
        total_val = total_cost = 0.0
        for ticker, info in ticker_info.items():
            try:
                price = float(price_df.loc[idx, ticker])
                if pd.isna(price) or price <= 0: continue
            except (KeyError, TypeError, ValueError): continue
            shares = cost = 0.0
            for lot in info["lots"]:
                if (lot.get("date") or "2000-01-01")[:10] <= date_str:
                    s = float(lot.get("shares", 0))
                    shares += s
                    cost   += s * float(lot.get("price", 0))
            if shares > 0:
                to_sgd     = info["to_sgd"]
                total_val  += shares * price * to_sgd
                total_cost += cost   * to_sgd
        if total_val > 0 or total_cost > 0:
            dates_out.append(date_str)
            portfolio_out.append(round(total_val, 2))
            cost_out.append(round(total_cost, 2))
    sectors: dict = {}
    for h in holdings:
        sector  = h.get("sector") or "Other"
        to_sgd  = _to_sgd_rate((h.get("currency") or "USD").upper(), fx)
        price   = h.get("current_price") or h.get("cost_price", 0)
        mkt_sgd = (h.get("shares", 0) * price) * to_sgd
        sectors[sector] = round(sectors.get(sector, 0) + mkt_sgd, 2)
    # ── Benchmark overlay (SPY, normalised to portfolio start value) ──
    benchmark = request.args.get("benchmark", "").upper().strip()
    bench_out = []
    if benchmark and dates_out:
        try:
            bdf = yf.download(benchmark, period=period, auto_adjust=True, progress=False)
            if not bdf.empty:
                import pandas as pd
                bcl = bdf["Close"].resample("W").last().ffill()
                start_bench = None
                start_port  = portfolio_out[0] if portfolio_out else None
                for idx in price_df.index:
                    date_str = idx.strftime("%Y-%m-%d")
                    if date_str not in dates_out:
                        continue
                    try:
                        bv = float(bcl.loc[idx])
                        if pd.isna(bv): continue
                    except (KeyError, TypeError):
                        continue
                    if start_bench is None:
                        start_bench = bv
                    # Scale so benchmark starts at same value as portfolio
                    scaled = round(bv / start_bench * start_port, 2) if start_bench and start_port else bv
                    bench_out.append(scaled)
                # Pad / trim to same length as dates_out
                while len(bench_out) < len(dates_out): bench_out.append(None)
                bench_out = bench_out[:len(dates_out)]
        except Exception:
            bench_out = []

    result = {
        "dates": dates_out, "portfolio_sgd": portfolio_out, "cost_sgd": cost_out,
        "sectors": dict(sorted(sectors.items(), key=lambda x: -x[1])),
        "benchmark": bench_out, "benchmark_ticker": benchmark or None,
    }
    _perf_cache[cache_key] = {"data": result, "ts": datetime.now()}
    return jsonify(result)



# ─── ANALYTICS ───────────────────────────────────────────────────────────────

@app.route("/api/analytics/portfolio")
def analytics_portfolio():
    fx       = get_fx_rate("USD", "SGD")
    equities = list_equities()
    enriched = [_enrich_equity(e, fx) for e in equities]
    crypto   = list_crypto()
    trades   = list_trades()

    # ── Sector breakdown (SGD) ──
    sectors: dict = {}
    for e in enriched:
        s = e.get("sector") or "Other"
        sectors[s] = round(sectors.get(s, 0) + (e.get("market_value_sgd") or 0), 2)

    # ── Currency exposure (SGD equivalent) ──
    currencies: dict = {}
    for e in enriched:
        ccy = e.get("currency") or "USD"
        currencies[ccy] = round(currencies.get(ccy, 0) + (e.get("market_value_sgd") or 0), 2)
    cr_val_usd = sum((c.get("current_price") or 0) * (c.get("quantity") or 0) for c in crypto)
    if cr_val_usd > 0:
        currencies["CRYPTO"] = round(currencies.get("CRYPTO", 0) + cr_val_usd * fx, 2)

    # ── Platform breakdown ──
    platforms: dict = {}
    for e in enriched:
        p = e.get("platform") or "Unknown"
        platforms[p] = round(platforms.get(p, 0) + (e.get("market_value_sgd") or 0), 2)

    # ── Top holdings ──
    top = sorted(enriched, key=lambda e: e.get("market_value_sgd") or 0, reverse=True)[:10]

    # ── Risk metrics ──
    total_eq_sgd = sum(e.get("market_value_sgd") or 0 for e in enriched)
    total_cr_sgd = cr_val_usd * fx
    total_sgd    = total_eq_sgd + total_cr_sgd

    hhi = max_w = 0.0
    if total_sgd > 0:
        weights = [(e.get("market_value_sgd") or 0) / total_sgd for e in enriched]
        # add crypto as one block
        if total_cr_sgd > 0:
            weights.append(total_cr_sgd / total_sgd)
        hhi   = round(sum(w * w for w in weights), 4)
        max_w = round(max(weights) * 100, 1) if weights else 0

    # Diversification score 0-100 (inverse of HHI, scaled)
    div_score = round(max(0, min(100, (1 - hhi) * 100)), 1)

    # Unrealised P&L totals
    total_pnl_sgd  = sum(e.get("unrealized_pnl_sgd") or 0 for e in enriched)
    total_cost_sgd = sum(e.get("cost_basis_sgd") or 0 for e in enriched)
    total_pnl_pct  = round(total_pnl_sgd / total_cost_sgd * 100, 2) if total_cost_sgd else 0

    # Options summary
    open_opts   = [t for t in trades if (t.get("status") or "").upper() == "OPEN"]
    closed_opts = [t for t in trades if (t.get("status") or "").upper() == "CLOSED"]
    opt_pnl     = sum(float(t.get("pnl") or 0) for t in closed_opts)
    wins        = sum(1 for t in closed_opts if float(t.get("pnl") or 0) > 0)
    win_rate    = round(wins / len(closed_opts) * 100, 1) if closed_opts else None

    return jsonify({
        "fx_usdsgd"  : fx,
        "sectors"    : dict(sorted(sectors.items(),   key=lambda x: -x[1])),
        "currencies" : dict(sorted(currencies.items(), key=lambda x: -x[1])),
        "platforms"  : dict(sorted(platforms.items(),  key=lambda x: -x[1])),
        "top_holdings": [
            {k: v for k, v in e.items()
             if k in ("id","ticker","company","sector","currency","platform",
                      "shares","cost_price","current_price",
                      "market_value_sgd","cost_basis_sgd",
                      "unrealized_pnl_sgd","unrealized_pnl_pct")}
            for e in top
        ],
        "risk": {
            "hhi"           : hhi,
            "max_weight_pct": max_w,
            "div_score"     : div_score,
            "num_holdings"  : len(enriched),
            "num_sectors"   : len(sectors),
            "total_sgd"     : round(total_sgd, 2),
            "total_eq_sgd"  : round(total_eq_sgd, 2),
            "total_cr_sgd"  : round(total_cr_sgd, 2),
            "total_pnl_sgd" : round(total_pnl_sgd, 2),
            "total_pnl_pct" : total_pnl_pct,
            "open_options"  : len(open_opts),
            "opt_pnl"       : round(opt_pnl, 2),
            "win_rate"      : win_rate,
        },
    })


# ─── WATCHLIST ────────────────────────────────────────────────────────────────

@app.route("/api/watchlist", methods=["GET"])
def watchlist_list():
    return jsonify(list_watchlist())

@app.route("/api/watchlist", methods=["POST"])
def watchlist_add():
    body   = request.get_json(force=True, silent=True) or {}
    ticker = (body.get("ticker") or "").strip().upper()
    if not ticker:
        return jsonify({"error": "ticker required"}), 400
    new_id = add_to_watchlist(ticker, body.get("sector"), body.get("notes"), body.get("score"), body.get("screener_theme"))
    return jsonify({"id": new_id, "ticker": ticker}), 201

@app.route("/api/watchlist/<ticker>", methods=["DELETE"])
def watchlist_remove(ticker):
    remove_from_watchlist(ticker)
    return jsonify({"status": "removed"})


# ─── ERROR HANDLERS ───────────────────────────────────────────────────────────

# --- PDF REPORT ---

@app.route("/api/report/pdf")
def report_pdf():
    if not PDF_AVAILABLE:
        return jsonify({"error": "reportlab/pypdf not installed"}), 503
    try:
        fx       = get_fx_rate("USD", "SGD")
        equities = list_equities()
        enriched = [_enrich_equity(e, fx) for e in equities]
        crypto   = list_crypto()
        trades   = list_trades()

        sectors = {}
        for e in enriched:
            s = e.get("sector") or "Unknown"
            sectors[s] = sectors.get(s, 0) + (e.get("market_value_sgd") or 0)

        currencies = {}
        for e in enriched:
            c = e.get("currency") or "USD"
            currencies[c] = currencies.get(c, 0) + (e.get("market_value_sgd") or 0)
        for cr in crypto:
            sgd_val = float(cr.get("quantity") or 0) * float(cr.get("current_price") or 0) * fx
            currencies["CRYPTO"] = currencies.get("CRYPTO", 0) + sgd_val

        platforms = {}
        for e in enriched:
            p = e.get("platform") or "Unknown"
            platforms[p] = platforms.get(p, 0) + (e.get("market_value_sgd") or 0)

        top = sorted(enriched, key=lambda e: e.get("market_value_sgd") or 0, reverse=True)[:10]

        total_eq_sgd  = sum(e.get("market_value_sgd") or 0 for e in enriched)
        total_cr_sgd  = currencies.get("CRYPTO", 0)
        total_sgd     = total_eq_sgd + total_cr_sgd
        total_pnl     = sum(e.get("unrealized_pnl_sgd") or 0 for e in enriched)
        cost_basis    = total_sgd - total_pnl
        total_pnl_pct = round((total_pnl / cost_basis * 100) if cost_basis else 0, 2)

        weights   = [(e.get("market_value_sgd") or 0) / total_sgd for e in enriched] if total_sgd else []
        hhi       = round(sum(w * w for w in weights), 4) if weights else 0
        div_score = round(max(0, min(100, (1 - hhi) * 100)), 1)
        max_w     = round(max((w * 100 for w in weights), default=0), 1)

        open_opts   = [t for t in trades if not t.get("close_date")]
        closed_opts = [t for t in trades if t.get("close_date")]
        opt_pnl     = sum(float(t.get("pnl") or 0) for t in closed_opts)
        winners     = [t for t in closed_opts if float(t.get("pnl") or 0) > 0]
        win_rate    = round(len(winners) / len(closed_opts) * 100, 1) if closed_opts else None

        data = {
            "fx_usdsgd"   : fx,
            "sectors"     : {k: round(v, 2) for k, v in sorted(sectors.items(),    key=lambda x: -x[1])},
            "currencies"  : {k: round(v, 2) for k, v in sorted(currencies.items(), key=lambda x: -x[1])},
            "platforms"   : {k: round(v, 2) for k, v in sorted(platforms.items(),  key=lambda x: -x[1])},
            "top_holdings": top,
            "open_trades" : open_opts,
            "risk": {
                "hhi"           : hhi,
                "max_weight_pct": max_w,
                "div_score"     : div_score,
                "num_holdings"  : len(enriched),
                "num_sectors"   : len(sectors),
                "total_sgd"     : round(total_sgd, 2),
                "total_eq_sgd"  : round(total_eq_sgd, 2),
                "total_cr_sgd"  : round(total_cr_sgd, 2),
                "total_pnl_sgd" : round(total_pnl, 2),
                "total_pnl_pct" : total_pnl_pct,
                "open_options"  : len(open_opts),
                "opt_pnl"       : round(opt_pnl, 2),
                "win_rate"      : win_rate,
            },
        }

        pdf_bytes = build_portfolio_pdf(data)
        from datetime import datetime as _dt
        fname = "portfolio_report_{}.pdf".format(_dt.now().strftime("%Y%m%d_%H%M"))
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=\"{}\"".format(fname)},
        )
    except Exception as exc:
        import traceback
        return jsonify({"error": str(exc), "trace": traceback.format_exc()}), 500


# ─── EARNINGS CALENDAR ───────────────────────────────────────────────────────

_earnings_cache: dict = {}
_EARNINGS_TTL = 6 * 3600   # refresh every 6 hours (earnings dates change rarely)

@app.route("/api/earnings")
def earnings_calendar():
    """Return upcoming earnings dates for all portfolio tickers within next 90 days."""
    from datetime import date, timedelta, datetime as _dt
    import pandas as pd

    today  = date.today()
    cutoff = today + timedelta(days=90)

    tickers: set = set()
    for eq in list_equities():
        t = (eq.get("ticker") or "").strip().upper()
        if t:
            tickers.add(t)
    for trade in list_trades(status="OPEN"):
        t = (trade.get("ticker") or "").strip().upper()
        if t:
            tickers.add(t)
    for wl in list_watchlist():
        t = (wl.get("ticker") or "").strip().upper()
        if t:
            tickers.add(t)

    results = []
    for ticker in sorted(tickers):
        cached = _earnings_cache.get(ticker)
        if cached and (_dt.now() - cached["ts"]).total_seconds() < _EARNINGS_TTL:
            entry = cached["entry"]
            if entry:
                try:
                    entry_date = _dt.strptime(entry["date"], "%Y-%m-%d").date()
                    if today <= entry_date <= cutoff:
                        results.append(entry)
                except Exception:
                    pass
            continue

        try:
            cal = yf.Ticker(ticker).calendar
            if cal is None:
                _earnings_cache[ticker] = {"entry": None, "ts": _dt.now()}
                continue

            earn_dates = []
            eps_est    = None

            if isinstance(cal, dict):
                raw_dates = cal.get("Earnings Date") or cal.get("earnings_date") or []
                if not isinstance(raw_dates, list):
                    raw_dates = [raw_dates]
                earn_dates = raw_dates
                eps_est    = cal.get("Earnings Average") or cal.get("EPS Estimate")
            elif isinstance(cal, pd.DataFrame):
                try:
                    earn_dates = [pd.Timestamp(str(cal.loc["Earnings Date"].iloc[0]))]
                    eps_est    = cal.loc["Earnings Average"].iloc[0] if "Earnings Average" in cal.index else None
                except Exception:
                    earn_dates = []

            entry = None
            for ed in earn_dates:
                if ed is None:
                    continue
                try:
                    if hasattr(ed, "date"):
                        ed_date = ed.date()
                    else:
                        ed_date = _dt.strptime(str(ed)[:10], "%Y-%m-%d").date()

                    if today <= ed_date <= cutoff:
                        when = "—"
                        if hasattr(ed, "hour"):
                            when = "BMO" if ed.hour < 12 else "AMC"

                        entry = {
                            "ticker"      : ticker,
                            "date"        : str(ed_date),
                            "days_away"   : (ed_date - today).days,
                            "eps_estimate": round(float(eps_est), 2) if eps_est is not None else None,
                            "when"        : when,
                        }
                        results.append(entry)
                        break
                except Exception:
                    continue

            _earnings_cache[ticker] = {"entry": entry, "ts": _dt.now()}

        except Exception:
            _earnings_cache[ticker] = {"entry": None, "ts": _dt.now()}

    results.sort(key=lambda x: x["date"])
    return jsonify({"earnings": results, "as_of": str(today)})



# --- AI ASSISTANT ---

try:
    import anthropic as _anthropic
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _ANTHROPIC_AVAILABLE = False

_ASSISTANT_SYSTEM = (
    "You are an AI investment assistant built into a personal portfolio management system.\n"
    "You have access to the user's live portfolio data provided below. Be concise, direct, and helpful.\n"
    "Focus on actionable insights about their specific positions.\n"
    "Format numbers clearly (e.g. $1,234.56, +12.3%). Keep responses under 200 words unless the user asks for detail.\n"
    "Do not give generic financial advice disclaimers - the user is an experienced investor who understands the risks.\n"
    "Today's date: {date}\n\n"
    "=== LIVE PORTFOLIO CONTEXT ===\n"
    "{context}"
)


def _build_portfolio_context():
    """Assemble a concise portfolio snapshot to inject into the assistant prompt."""
    lines = []
    try:
        fx  = get_fx_rate("USD", "SGD")
        eqs = [_enrich_equity(e, fx) for e in list_equities()]
        if eqs:
            total_val = sum(e.get("market_value_sgd", 0) for e in eqs)
            total_pnl = sum(e.get("unrealized_pnl_sgd", 0) for e in eqs)
            lines.append(f"EQUITIES - {len(eqs)} holdings, total S${total_val:,.0f}, unrealised P&L S${total_pnl:+,.0f}")
            for e in sorted(eqs, key=lambda x: -(x.get("market_value_sgd") or 0))[:8]:
                lines.append(
                    f"  {e['ticker']} {e.get('shares',0):.0f}sh @ {e.get('current_price','?')} "
                    f"| cost {e.get('cost_price','?')} | P&L S${e.get('unrealized_pnl_sgd',0):+,.0f} "
                    f"({e.get('unrealized_pnl_pct',0):+.1f}%)"
                )
    except Exception:
        pass
    try:
        open_opts = list_trades(status="OPEN")
        if open_opts:
            lines.append(f"\nOPEN OPTIONS - {len(open_opts)} legs")
            seen = set()
            for t in open_opts:
                key = t.get("trade_group") or t.get("id")
                if key in seen:
                    continue
                seen.add(key)
                lines.append(
                    f"  {t['ticker']} {t.get('strategy','')} | "
                    f"{t.get('direction','')} {t.get('option_type','')} "
                    f"${t.get('strike','')} exp {str(t.get('expiry',''))[:10]} "
                    f"x{t.get('contracts',1)} @ ${t.get('premium',0)}"
                )
    except Exception:
        pass
    try:
        closed = list_trades(status="CLOSED")
        if closed:
            total_pnl = sum(float(t.get("pnl") or 0) for t in closed)
            wins = sum(1 for t in closed if float(t.get("pnl") or 0) > 0)
            win_rate = wins / len(closed) * 100
            lines.append(
                f"\nCLOSED OPTIONS - {len(closed)} trades | "
                f"total P&L ${total_pnl:+.2f} | win rate {win_rate:.0f}%"
            )
    except Exception:
        pass
    try:
        wl = list_watchlist()
        if wl:
            tickers = ", ".join(w["ticker"] for w in wl[:10])
            lines.append(f"\nWATCHLIST - {tickers}")
    except Exception:
        pass
    return "\n".join(lines) if lines else "No portfolio data available yet."


@app.route("/api/assistant", methods=["POST"])
def assistant():
    if not _ANTHROPIC_AVAILABLE:
        return jsonify({"error": "anthropic package not installed. Run: pip install anthropic"}), 503
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY environment variable not set"}), 503

    if DEMO_MODE:
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
        if not _assistant_quota_ok(ip):
            return jsonify({"error": f"Demo limit reached — this public demo allows {DEMO_ASSISTANT_CAP} AI assistant messages per visitor per day."}), 429
        _assistant_quota_hit(ip)

    body    = request.get_json() or {}
    query   = (body.get("query") or "").strip()
    history = body.get("history") or []

    if not query:
        return jsonify({"error": "query is required"}), 400

    from datetime import date
    context = _build_portfolio_context()
    system  = _ASSISTANT_SYSTEM.format(date=date.today().isoformat(), context=context)

    messages = []
    for h in history[-10:]:
        if h.get("role") in ("user", "assistant") and h.get("content"):
            messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": query})

    try:
        client   = _anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model      = "claude-haiku-4-5-20251001",
            max_tokens = 512,
            system     = system,
            messages   = messages,
        )
        answer = response.content[0].text
        return jsonify({"answer": answer, "tokens": response.usage.output_tokens})
    except Exception as e:
        return jsonify({"error": str(e)}), 502


# ─── NET WORTH ────────────────────────────────────────────────────────────────

@app.route("/api/networth", methods=["GET"])
def get_networth():
    usd_sgd = get_fx_rate("USD", "SGD")
    items   = list_networth_items()

    def to_sgd(amount, currency):
        return round(float(amount) * _to_sgd_rate((currency or "SGD").upper(), usd_sgd), 2)

    CATEGORY_ORDER = ["cash", "cpf_srs", "property", "insurance", "investment", "other", "liability"]

    category_map: dict  = {}
    total_assets        = 0.0
    total_liabilities   = 0.0

    for it in items:
        sgd = to_sgd(it["amount"], it.get("currency", "SGD"))
        it["sgd_value"] = sgd
        cat = it.get("category", "other")
        if cat not in category_map:
            category_map[cat] = {"items": [], "subtotal": 0.0}
        category_map[cat]["items"].append(it)
        if cat == "liability":
            total_liabilities += sgd
        else:
            total_assets += sgd

    for cat in category_map.values():
        cat["subtotal"] = round(sum(i["sgd_value"] for i in cat["items"]), 2)

    ordered_cats = {k: category_map[k] for k in CATEGORY_ORDER if k in category_map}
    for k in category_map:
        if k not in ordered_cats:
            ordered_cats[k] = category_map[k]

    return jsonify({
        "items":      items,
        "categories": ordered_cats,
        "summary": {
            "total_assets":      round(total_assets, 2),
            "total_liabilities": round(total_liabilities, 2),
            "net_worth":         round(total_assets - total_liabilities, 2),
        },
        "fx_usdsgd": usd_sgd,
    })


@app.route("/api/networth", methods=["POST"])
def create_networth_item():
    data   = request.get_json() or {}
    new_id = add_networth_item(data)
    return jsonify({"id": new_id, "ok": True}), 201


@app.route("/api/networth/<int:item_id>", methods=["PUT"])
def edit_networth_item(item_id):
    data = request.get_json() or {}
    update_networth_item(item_id, data)
    return jsonify({"ok": True})


@app.route("/api/networth/<int:item_id>", methods=["DELETE"])
def remove_networth_item(item_id):
    delete_networth_item(item_id)
    return jsonify({"ok": True})


# --- STOCK ANALYSER ---

@app.route("/api/analyse/<ticker>")
def analyse_stock(ticker):
    import math
    try:
        t   = yf.Ticker(ticker.upper())
        info = t.info or {}
        if not info.get("regularMarketPrice") and not info.get("currentPrice"):
            return jsonify({"error": "Ticker not found"}), 404

        price      = float(info.get("regularMarketPrice") or info.get("currentPrice") or 0)
        prev_close = float(info.get("previousClose") or price)
        change     = round(price - prev_close, 4)
        change_pct = round((change / prev_close * 100) if prev_close else 0, 4)
        shares     = float(info.get("sharesOutstanding") or 1)

        # ── Financials ──────────────────────────────────────────────

        # ── helpers ─────────────────────────────────────────────────
        import math

        def safe(df, row, col=0):
            try:
                if df is None or df.empty or row not in df.index: return None
                v = df.loc[row].iloc[col]
                return None if (v is None or (isinstance(v, float) and math.isnan(v))) else float(v)
            except:
                return None

        def df_to_records(df):
            if df is None or df.empty: return []
            cols = [str(c.date()) if hasattr(c, 'date') else str(c) for c in df.columns]
            rows = []
            for idx in df.index:
                row = {"label": str(idx)}
                for i, c in enumerate(cols):
                    v = df.loc[idx].iloc[i]
                    row[c] = None if (v is None or (isinstance(v, float) and math.isnan(v))) else float(v)
                rows.append(row)
            return rows

        try: fin = t.financials
        except: fin = None
        try: bs  = t.balance_sheet
        except: bs = None
        try: cf_stmt = t.cashflow
        except: cf_stmt = None

        revenue    = safe(fin, "Total Revenue")
        net_income = safe(fin, "Net Income")
        # Operating CF (not FCF) matches StockOracle's DCF base
        op_cf      = safe(cf_stmt, "Operating Cash Flow") or safe(cf_stmt, "Total Cash From Operating Activities")
        capex      = safe(cf_stmt, "Capital Expenditure") or safe(cf_stmt, "Purchase Of Property Plant And Equipment")
        fcf_raw    = (op_cf or 0) + (capex or 0)
        op_cf_per_sh  = op_cf / shares if op_cf and shares else None
        fcf_per_sh    = fcf_raw / shares if fcf_raw and shares else None
        ni_per_sh     = net_income / shares if net_income and shares else None

        # ── intrinsic value inputs ────────────────────────────────────
        eps        = float(info.get("trailingEps")  or 0)
        eps_fwd    = float(info.get("forwardEps")   or eps)
        bv_per_sh  = float(info.get("bookValue")    or 0)
        rev_per_sh = float(info.get("revenuePerShare") or (revenue / shares if revenue else 0))
        pe         = float(info.get("trailingPE")   or 0)
        ps         = float(info.get("priceToSalesTrailing12Months") or 0)
        pb         = float(info.get("priceToBook")  or 0)

        # Use revenue growth as DCF growth rate — more stable than earnings growth
        # (earningsGrowth can spike to 100%+ for a single exceptional year)
        rev_g     = float(info.get("revenueGrowth") or 0.15)
        raw_eg    = float(info.get("earningsGrowth") or 0)
        dcf_g     = max(0.03, min(0.40, rev_g if rev_g > 0 else 0.10))
        # g_rate keeps earnings growth for scoring/other use
        g_rate    = max(0.05, min(0.40, raw_eg if raw_eg > 0 else rev_g))

        # ── DCF: operating CF/share as base, revenue growth as rate ──
        # StockOracle uses operating CF (not FCF) and revenue growth rate.
        # Years 1-10: dcf_g; years 11-20: dcf_g * 0.5; terminal: 3%.
        def dcf(cf_ps, gr, years=20, dr=0.10, tgr=0.03):
            if not cf_ps or cf_ps <= 0: return None
            pv, c = 0.0, cf_ps
            for yr in range(1, years + 1):
                g = gr if yr <= 10 else gr * 0.5
                c = c * (1 + g)
                pv += c / (1 + dr) ** yr
            tv  = c * (1 + tgr) / (dr - tgr)
            pv += tv / (1 + dr) ** years
            return round(pv, 2)

        def dcf_terminal(cf_ps, gr, years=10, dr=0.10, tgr=0.03):
            if not cf_ps or cf_ps <= 0: return None
            pv, c = 0.0, cf_ps
            for yr in range(1, years + 1):
                c = c * (1 + gr)
                pv += c / (1 + dr) ** yr
            tv  = c * (1 + tgr) / (dr - tgr)
            pv += tv / (1 + dr) ** years
            return round(pv, 2)

        # ── Historical median multiples from 4-year yfinance data ────
        # StockOracle uses the stock's own historical median P/E, P/S, P/B
        # rather than today's (possibly compressed) multiples.
        hist_pe = pe; hist_ps = ps; hist_pb = pb
        try:
            hist_prices = t.history(period="4y", interval="3mo")["Close"]
            fin_cols = list(fin.columns) if fin is not None and not fin.empty else []
            if hist_prices is not None and not hist_prices.empty and fin_cols:
                pe_list, ps_list, pb_list = [], [], []
                for col in fin_cols:
                    yr = col.year if hasattr(col, "year") else None
                    if not yr: continue
                    ni_yr  = safe(fin, "Net Income",  list(fin.columns).index(col))
                    rev_yr = safe(fin, "Total Revenue", list(fin.columns).index(col))
                    eq_yr  = safe(bs,  "Stockholders Equity", list(bs.columns).index(col)) if bs is not None and not bs.empty and "Stockholders Equity" in bs.index else None
                    # Average price for that calendar year
                    yr_prices = hist_prices[hist_prices.index.year == yr]
                    if yr_prices.empty: continue
                    avg_price = float(yr_prices.mean())
                    if ni_yr and ni_yr > 0:
                        pe_list.append(avg_price / (ni_yr / shares))
                    if rev_yr and rev_yr > 0:
                        ps_list.append(avg_price / (rev_yr / shares))
                    if eq_yr and eq_yr > 0:
                        pb_list.append(avg_price / (eq_yr / shares))
                if pe_list:
                    import statistics
                    hist_pe = statistics.median(pe_list)
                if ps_list:
                    hist_ps = statistics.median(ps_list)
                if pb_list:
                    hist_pb = statistics.median(pb_list)
        except Exception:
            pass  # fall back to current multiples

        mean_pe_val  = round(hist_pe * eps, 2)       if hist_pe > 0 and eps > 0 else None
        mean_ps_val  = round(hist_ps * rev_per_sh, 2) if hist_ps > 0 and rev_per_sh > 0 else None
        mean_pb_val  = round(hist_pb * bv_per_sh, 2)  if hist_pb > 0 and bv_per_sh > 0 else None

        # ── PEG: EPS × earnings_growth%  (Lynch model, uncapped) ────
        peg_g   = raw_eg if raw_eg > 0 else dcf_g
        peg_val = round(eps * (peg_g * 100), 2) if eps > 0 and peg_g > 0 else None

        # ── PSG: EPS × revenue_growth%  (revenue-growth PEG variant) ─
        psg_val = round(eps * (rev_g * 100), 2) if eps > 0 and rev_g > 0 else None

        graham_val = round((22.5 * eps * bv_per_sh) ** 0.5, 2) if eps > 0 and bv_per_sh > 0 else None

        iv = {
            "dcf_20":        dcf(op_cf_per_sh, dcf_g, years=20),
            "dfcf_20":       dcf(op_cf_per_sh * 0.8 if op_cf_per_sh else None, dcf_g, years=20),
            "dni_20":        dcf(ni_per_sh, dcf_g, years=20),
            "dfcf_terminal": dcf_terminal(op_cf_per_sh * 0.8 if op_cf_per_sh else None, dcf_g),
            "mean_pe":       mean_pe_val,
            "mean_ps":       mean_ps_val,
            "mean_pb":       mean_pb_val,
            "peg":           peg_val,
            "psg":           psg_val,
            "graham":        graham_val,
        }
        # expose inputs for frontend debug panel
        iv["_inputs"] = {
            "op_cf_per_sh": round(op_cf_per_sh, 4) if op_cf_per_sh else None,
            "fcf_per_sh":   round(fcf_per_sh, 4) if fcf_per_sh else None,
            "ni_per_sh":    round(ni_per_sh, 4) if ni_per_sh else None,
            "dcf_g_pct":    round(dcf_g * 100, 2),
            "peg_g_pct":    round(peg_g * 100, 2),
            "rev_g_pct":    round(rev_g * 100, 2),
            "hist_pe":      round(hist_pe, 2),
            "hist_ps":      round(hist_ps, 2),
            "hist_pb":      round(hist_pb, 2),
        }
        weights = {"dcf_20": 0.15, "dfcf_20": 0.12, "dni_20": 0.10,
                   "dfcf_terminal": 0.13, "mean_pe": 0.15, "mean_ps": 0.10,
                   "mean_pb": 0.10, "peg": 0.08, "psg": 0.07}
        tot_w, composite = 0.0, 0.0
        for k, w in weights.items():
            v = iv.get(k)
            if v and v > 0:
                composite += v * w; tot_w += w
        iv["composite"] = round(composite / tot_w, 2) if tot_w > 0 else None

        # ── scores ───────────────────────────────────────────────────
        def clamp(v): return min(max(round(v, 1), 1), 10)

        def score_profitability():
            s = 4.0
            gm  = float(info.get("grossMargins")  or 0)
            npm = float(info.get("profitMargins")  or 0)
            roe = float(info.get("returnOnEquity") or 0)
            roa = float(info.get("returnOnAssets") or 0)
            if gm  > 0.40: s += 1
            if gm  > 0.60: s += 1
            if npm > 0.10: s += 1
            if npm > 0.20: s += 0.5
            if roe > 0.15: s += 0.5
            if roe > 0.30: s += 0.5
            if roa > 0.10: s += 0.5
            return clamp(s)

        def score_growth():
            s = 4.0
            rg = float(info.get("revenueGrowth")  or 0)
            eg = float(info.get("earningsGrowth") or 0)
            if rg > 0.10: s += 1
            if rg > 0.25: s += 1
            if eg > 0.10: s += 1
            if eg > 0.30: s += 1
            if eps_fwd > eps and eps > 0: s += 0.5
            return clamp(s)

        def score_fin_strength():
            s = 4.0
            cr = float(info.get("currentRatio") or 0)
            de = float(info.get("debtToEquity") or 999)
            if cr > 1.5: s += 1
            if cr > 2.5: s += 0.5
            if de < 100: s += 1
            if de <  50: s += 0.5
            if de <  20: s += 0.5
            if float(info.get("freeCashflow") or 0) > 0: s += 1
            return clamp(s)

        def score_valuation():
            s = 5.0
            pe_v  = float(info.get("trailingPE") or 999)
            pb_v  = float(info.get("priceToBook") or 999)
            ps_v  = float(info.get("priceToSalesTrailing12Months") or 999)
            peg_v = float(info.get("trailingPegRatio") or 999)
            if pe_v < 15: s += 2
            elif pe_v < 25: s += 1
            elif pe_v > 60: s -= 1
            if pb_v < 3:  s += 0.5
            if ps_v < 5:  s += 0.5
            if 0 < peg_v < 1: s += 1
            elif 1 <= peg_v < 2: s += 0.5
            if iv.get("composite") and iv["composite"] > price: s += 1
            return clamp(s)

        def score_dividend():
            dy = float(info.get("dividendYield") or 0)
            pr = float(info.get("payoutRatio")   or 0)
            if dy <= 0: return 1
            s = 4.0
            if dy > 0.01: s += 1
            if dy > 0.02: s += 1
            if dy > 0.03: s += 1
            if dy > 0.04: s += 1
            if 0 < pr < 0.6: s += 1
            if pr > 0.9: s -= 2
            return clamp(s)

        # ── Multi-year chart data ─────────────────────────────────────
        import statistics as _stats
        chart_data = {"income": {}, "cashflow": {}, "balance": {}, "ccc": {}, "receivables": {}}
        _rev_history = []  # [(year_int, revenue)] for predictability

        try:
            if fin is not None and not fin.empty:
                for col in sorted(fin.columns, key=lambda c: c.year if hasattr(c,'year') else 0):
                    yr  = str(col.year) if hasattr(col,'year') else str(col)[:4]
                    idx = list(fin.columns).index(col)
                    rev = safe(fin, "Total Revenue", idx)
                    oi  = (safe(fin, "Operating Income", idx)
                           or safe(fin, "Total Operating Income As Reported", idx)
                           or safe(fin, "EBIT", idx))
                    ni  = (safe(fin, "Net Income", idx)
                           or safe(fin, "Net Income Common Stockholders", idx)
                           or safe(fin, "Net Income Including Noncontrolling Interests", idx))
                    chart_data["income"][yr] = {"revenue": rev, "operating_income": oi, "net_income": ni}
                    if rev and rev > 0:
                        _rev_history.append((int(yr), rev))
        except Exception: pass

        try:
            if cf_stmt is not None and not cf_stmt.empty:
                for col in sorted(cf_stmt.columns, key=lambda c: c.year if hasattr(c,'year') else 0):
                    yr  = str(col.year) if hasattr(col,'year') else str(col)[:4]
                    idx = list(cf_stmt.columns).index(col)
                    op_c = safe(cf_stmt, "Operating Cash Flow", idx) or safe(cf_stmt, "Total Cash From Operating Activities", idx)
                    cap  = safe(cf_stmt, "Capital Expenditure", idx)
                    sbc  = (safe(cf_stmt, "Stock Based Compensation", idx)
                             or safe(cf_stmt, "Share Based Compensation", idx))
                    fcf_y = (op_c or 0) + (cap or 0) if op_c is not None else None
                    ni_y  = chart_data["income"].get(yr, {}).get("net_income")
                    chart_data["cashflow"][yr] = {"op_cf": op_c, "fcf": fcf_y, "net_income": ni_y, "sbc": sbc}
        except Exception: pass

        try:
            if bs is not None and not bs.empty:
                for col in sorted(bs.columns, key=lambda c: c.year if hasattr(c,'year') else 0):
                    yr  = str(col.year) if hasattr(col,'year') else str(col)[:4]
                    idx = list(bs.columns).index(col)
                    cash_y = safe(bs, "Cash And Cash Equivalents", idx)
                    sti_y  = safe(bs, "Short Term Investments", idx) or safe(bs, "Other Short Term Investments", idx) or 0
                    debt_y = safe(bs, "Total Debt", idx) or safe(bs, "Long Term Debt", idx)
                    ar_y   = safe(bs, "Net Receivables", idx) or safe(bs, "Accounts Receivable", idx)
                    inv_y  = safe(bs, "Inventory", idx)
                    ap_y   = safe(bs, "Accounts Payable", idx)
                    cash_tot = (cash_y or 0) + (sti_y or 0) if cash_y is not None else None
                    chart_data["balance"][yr] = {"cash": cash_tot, "total_debt": debt_y}
                    rev_y = chart_data["income"].get(yr, {}).get("revenue")
                    # COGS from income statement (Revenue - Gross Profit)
                    cogs_y = None
                    try:
                        fin_col_match = [c for c in fin.columns if (str(c.year) if hasattr(c,'year') else str(c)[:4]) == yr]
                        if fin_col_match and fin is not None:
                            fi = list(fin.columns).index(fin_col_match[0])
                            gp = safe(fin, "Gross Profit", fi)
                            if gp is not None and rev_y:
                                cogs_y = rev_y - gp
                    except Exception: pass
                    dso = round(ar_y  / rev_y  * 365, 1) if ar_y  and rev_y  and rev_y  > 0 else None
                    dio = round(inv_y / cogs_y * 365, 1) if inv_y and cogs_y and cogs_y > 0 else None
                    dpo = round(ap_y  / cogs_y * 365, 1) if ap_y  and cogs_y and cogs_y > 0 else None
                    ccc = round((dso or 0) + (dio or 0) - (dpo or 0), 1) if dso is not None else None
                    chart_data["ccc"][yr] = {"dso": dso, "dio": dio, "dpo": dpo, "ccc": ccc}
                    chart_data["receivables"][yr] = {"revenue": rev_y, "ar": ar_y}
        except Exception: pass

        shares_data = {
            "outstanding": info.get("sharesOutstanding"),
            "float":       info.get("floatShares"),
        }

        # ── Predictability score ──────────────────────────────────────
        def score_predictability():
            if len(_rev_history) < 3:
                return clamp(4.0)
            _sorted = sorted(_rev_history, key=lambda x: x[0])
            _vals   = [r[1] for r in _sorted]
            _growths = [(_vals[i]-_vals[i-1])/abs(_vals[i-1]) for i in range(1, len(_vals)) if _vals[i-1] != 0]
            if not _growths:
                return clamp(4.0)
            _mean = sum(_growths) / len(_growths)
            _std  = (_stats.stdev(_growths) if len(_growths) > 1 else 0)
            _cv   = _std / abs(_mean) if _mean != 0 else 10.0
            _neg  = sum(1 for g in _growths if g < 0)
            s = 9.0 - (_cv * 3.5) - (_neg * 1.5)
            return clamp(s)

        # ── Moat score ────────────────────────────────────────────────
        def score_moat():
            s = 3.0
            gm  = float(info.get("grossMargins")    or 0)
            npm = float(info.get("profitMargins")    or 0)
            roe = float(info.get("returnOnEquity")   or 0)
            opm = float(info.get("operatingMargins") or 0)
            if gm  > 0.60: s += 2.5
            elif gm  > 0.40: s += 1.5
            elif gm  > 0.25: s += 0.5
            if roe > 0.30: s += 2.0
            elif roe > 0.15: s += 1.0
            if npm > 0.20: s += 1.5
            elif npm > 0.10: s += 0.5
            if opm > 0.25: s += 0.5
            return clamp(s)

        scores = {
            "predictability":     score_predictability(),
            "profitability":      score_profitability(),
            "growth":             score_growth(),
            "moat":               score_moat(),
            "financial_strength": score_fin_strength(),
            "valuation":          score_valuation(),
        }

        hist_revenue = {}  # kept for backward compat
        try:
            if fin is not None and not fin.empty and "Total Revenue" in fin.index:
                for col in fin.columns:
                    yr  = str(col.year) if hasattr(col, "year") else str(col)[:4]
                    val = fin.loc["Total Revenue", col]
                    if val and not math.isnan(float(val)):
                        hist_revenue[yr] = float(val)
        except: pass

        # ── dividends ─────────────────────────────────────────────────
        dividends = []
        try:
            divs = t.dividends
            if divs is not None and not divs.empty:
                for dt, amt in divs.tail(20).items():
                    dividends.append({"date": str(dt.date()), "amount": round(float(amt), 4)})
        except: pass

        # ── news ──────────────────────────────────────────────────────
        news = []
        try:
            for n in (t.news or [])[:12]:
                news.append({
                    "title":     n.get("title", ""),
                    "url":       n.get("link", ""),
                    "publisher": n.get("publisher", ""),
                    "time":      n.get("providerPublishTime", 0),
                })
        except: pass

        # ── earnings ─────────────────────────────────────────────────
        earnings = []
        try:
            eh = t.earnings_history
            if eh is not None and not eh.empty:
                for _, row in eh.tail(8).iterrows():
                    earnings.append({
                        "period":       str(row.get("Quarter", "")),
                        "actual":       float(row.get("epsActual",    0) or 0),
                        "estimate":     float(row.get("epsEstimate",  0) or 0),
                        "surprise_pct": float(row.get("surprisePercent", 0) or 0),
                    })
        except: pass

        # ── helpers for output ────────────────────────────────────────
        def pct(v):
            try: return round(float(v) * 100, 2) if v is not None else None
            except: return None
        def fmt(v, d=2):
            try: return round(float(v), d) if v is not None else None
            except: return None

        return jsonify({
            "ticker": ticker.upper(),
            "info": {
                "name":        info.get("longName", ticker),
                "sector":      info.get("sector", ""),
                "industry":    info.get("industry", ""),
                "exchange":    info.get("exchange", ""),
                "currency":    info.get("currency", "USD"),
                "country":     info.get("country", ""),
                "employees":   info.get("fullTimeEmployees"),
                "description": info.get("longBusinessSummary", ""),
                "website":     info.get("website", ""),
            },
            "price": {
                "current":     fmt(price),
                "prev_close":  fmt(prev_close),
                "change":      fmt(change),
                "change_pct":  fmt(change_pct),
                "open":        fmt(info.get("open")),
                "day_high":    fmt(info.get("dayHigh")),
                "day_low":     fmt(info.get("dayLow")),
                "week52_high": fmt(info.get("fiftyTwoWeekHigh")),
                "week52_low":  fmt(info.get("fiftyTwoWeekLow")),
                "volume":      info.get("volume"),
                "avg_volume":  info.get("averageVolume"),
            },
            "metrics": {
                "market_cap":       info.get("marketCap"),
                "enterprise_value": info.get("enterpriseValue"),
                "pe_ratio":         fmt(info.get("trailingPE")),
                "forward_pe":       fmt(info.get("forwardPE")),
                "peg_ratio":        fmt(info.get("trailingPegRatio")),
                "ps_ratio":         fmt(info.get("priceToSalesTrailing12Months")),
                "pb_ratio":         fmt(info.get("priceToBook")),
                "ev_ebitda":        fmt(info.get("enterpriseToEbitda")),
                "ev_revenue":       fmt(info.get("enterpriseToRevenue")),
                "eps_ttm":          fmt(info.get("trailingEps")),
                "eps_forward":      fmt(info.get("forwardEps")),
                "revenue_ttm":      info.get("totalRevenue"),
                "gross_margin":     pct(info.get("grossMargins")),
                "operating_margin": pct(info.get("operatingMargins")),
                "net_margin":       pct(info.get("profitMargins")),
                "roe":              pct(info.get("returnOnEquity")),
                "roa":              pct(info.get("returnOnAssets")),
                "revenue_growth":   pct(info.get("revenueGrowth")),
                "earnings_growth":  pct(info.get("earningsGrowth")),
                "current_ratio":    fmt(info.get("currentRatio")),
                "quick_ratio":      fmt(info.get("quickRatio")),
                "debt_to_equity":   fmt(info.get("debtToEquity")),
                "total_debt":       info.get("totalDebt"),
                "total_cash":       info.get("totalCash"),
                "fcf":              info.get("freeCashflow"),
                "operating_cf":     info.get("operatingCashflow"),
                "beta":             fmt(info.get("beta")),
                "shares_outstanding": info.get("sharesOutstanding"),
                "shares_float":     info.get("floatShares"),
                "book_value":       fmt(info.get("bookValue")),
                "revenue_per_share":fmt(info.get("revenuePerShare")),
                "fcf_per_share":    fmt(fcf_per_sh),
                "dividend_rate":    fmt(info.get("dividendRate")),
                "dividend_yield":   pct(info.get("dividendYield")),
                "payout_ratio":     pct(info.get("payoutRatio")),
                "ex_div_date":      str(info.get("exDividendDate", "")),
                "next_earnings":    str(info.get("earningsDate", [""])[0] if isinstance(info.get("earningsDate"), list) else info.get("earningsDate", "")),
                "target_price":     fmt(info.get("targetMeanPrice")),
                "analyst_rating":   info.get("recommendationKey", "").upper(),
                "short_ratio":      fmt(info.get("shortRatio")),
            },
            "intrinsic_values": iv,
            "scores":       scores,
            "hist_revenue": hist_revenue,
            "chart_data":   chart_data,
            "shares_data":  shares_data,
            "fin_table":    df_to_records(fin),
            "bs_table":     df_to_records(bs),
            "cf_table":     df_to_records(cf_stmt),
            "dividends":    dividends,
            "news":         news,
            "earnings":     earnings,
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# --- MONTHLY SNAPSHOTS ---

@app.route("/api/snapshots/compute", methods=["POST"])
def compute_snapshot():
    """
    Fetch end-of-month prices for all holdings, compute net worth breakdown,
    save the snapshot, and return it.
    Body: { month: "YYYY-MM", fx_sgd: 0.74 }
    """
    from database import (get_conn, save_monthly_snapshot,
                          list_networth_items)
    import calendar as _cal
    data    = request.json or {}
    month   = data.get("month", "")          # e.g. "2026-05"
    if not month:
        return jsonify({"error": "month required"}), 400

    try:
        # ── End-of-month date ──────────────────────────────────────
        yr, mo   = int(month[:4]), int(month[5:7])
        last_day = _cal.monthrange(yr, mo)[1]
        target_date = datetime(yr, mo, last_day)
        eom_str     = target_date.date().isoformat()   # "2026-05-31"

        # ── Auto-fetch SGD/USD FX rate (same price_ref logic as equities) ──
        from datetime import timedelta, date as _date
        today     = datetime.now().date()
        price_ref = min(target_date.date(), today)
        fx_start  = (price_ref - timedelta(days=7)).strftime("%Y-%m-%d")
        fx_end    = (price_ref + timedelta(days=1)).strftime("%Y-%m-%d")
        sgd_per_usd = 1.35   # fallback
        fetched_fx_sgd = None
        try:
            fx_raw = yf.download("USDSGD=X", start=fx_start, end=fx_end,
                                 auto_adjust=True, progress=False, threads=False)
            if fx_raw is not None and not fx_raw.empty:
                fx_close = fx_raw["Close"] if "Close" in fx_raw else fx_raw
                if not fx_close.empty:
                    usdsgd = float(fx_close.dropna().iloc[-1])   # USD → SGD
                    sgd_per_usd = round(usdsgd, 4)
                    fetched_fx_sgd = round(1.0 / usdsgd, 4)     # SGD → USD (display)
        except Exception:
            pass
        app.logger.info(f"[compute] FX USDSGD={sgd_per_usd} (SGD/USD={fetched_fx_sgd})")

        def to_sgd(amount, ccy):
            c = (ccy or "SGD").upper()
            if c == "SGD": return float(amount or 0)
            if c == "USD": return float(amount or 0) * sgd_per_usd
            rate = get_fx_rate(c, "SGD")
            return float(amount or 0) * rate

        # ── Load equities filtered by purchase_date ────────────────
        with get_conn() as conn:
            equities = [dict(r) for r in conn.execute(
                """SELECT ticker, shares, currency, current_price, purchase_date
                   FROM equities
                   WHERE purchase_date IS NULL OR purchase_date <= ?
                   ORDER BY ticker""",
                (eom_str,)
            ).fetchall()]
            cryptos  = [dict(r) for r in conn.execute(
                "SELECT symbol, quantity, currency, current_price FROM crypto"
            ).fetchall()]
            cash_rows = [dict(r) for r in conn.execute(
                "SELECT amount, currency FROM cash"
            ).fetchall()] if conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='cash'"
            ).fetchone() else []

        app.logger.info(f"[compute] eom_str={eom_str} equities={len(equities)} cryptos={len(cryptos)}")
        if equities:
            app.logger.info(f"[compute] sample equity purchase_dates: { {e['ticker']: e['purchase_date'] for e in equities[:5]} }")

        # ── Fetch EOM equity prices (per-exchange) ───────────────
        # Group tickers by exchange — each has its own market calendar /
        # last trading day (US=NYSE/NASDAQ, SG=SGX, HK=HKEX).
        # Fetching groups separately avoids yfinance silently dropping
        # tickers when mixing exchanges in one batch call.
        tickers = list({e["ticker"] for e in equities})
        # price_ref already computed above (for FX fetch); reuse here
        price_end   = (price_ref + timedelta(days=1)).strftime("%Y-%m-%d")
        price_start = (price_ref - timedelta(days=7)).strftime("%Y-%m-%d")

        us_tickers = [t for t in tickers
                      if not t.endswith(".SI") and not t.endswith(".HK")]
        si_tickers = [t for t in tickers if t.endswith(".SI")]
        hk_tickers = [t for t in tickers if t.endswith(".HK")]

        prices     = {}
        price_date = eom_str   # updated below to the latest trading day seen

        def _fetch_group(group):
            nonlocal price_date
            if not group:
                return
            try:
                query = group[0] if len(group) == 1 else group
                app.logger.info(f"[compute] fetching {query!r} {price_start}→{price_end}")
                raw = yf.download(
                    query, start=price_start, end=price_end,
                    auto_adjust=True, progress=False, threads=False
                )
                app.logger.info(f"[compute] raw empty={getattr(raw,'empty',None)} shape={getattr(raw,'shape',None)}")
                if raw is None or (hasattr(raw, "empty") and raw.empty):
                    app.logger.warning(f"[compute] empty result for {group}")
                    return
                # Handle MultiIndex columns (yfinance >= 0.2 multi-ticker)
                if hasattr(raw.columns, "levels"):
                    close = raw["Close"]
                elif "Close" in raw.columns:
                    close = raw["Close"]
                else:
                    close = raw
                app.logger.info(f"[compute] close type={type(close).__name__} shape={getattr(close,'shape',None)} cols={list(getattr(close,'columns',[])) if hasattr(close,'columns') else 'series'}")
                if hasattr(close, "empty") and close.empty:
                    return
                last_date = str(close.index[-1])[:10]
                if last_date > price_date:
                    price_date = last_date
                if len(group) == 1:
                    # single ticker → close is a Series
                    series = close.dropna() if hasattr(close, "dropna") else close
                    if len(series):
                        val = float(series.iloc[-1])
                        prices[group[0]] = val
                        app.logger.info(f"[compute] {group[0]} = {val} ({last_date})")
                else:
                    # multiple tickers → close is DataFrame
                    for tk in group:
                        col = tk
                        if hasattr(close, "columns") and col in close.columns:
                            s = close[col].dropna()
                            if not s.empty:
                                val = float(s.iloc[-1])
                                prices[tk] = val
                                app.logger.info(f"[compute] {tk} = {val} ({last_date})")
                            else:
                                app.logger.warning(f"[compute] {tk} column all-NaN")
                        else:
                            app.logger.warning(f"[compute] {tk} not in close columns {list(getattr(close,'columns',[]))}")
            except Exception as exc:
                app.logger.error(f"[compute] _fetch_group({group}) error: {exc}", exc_info=True)

        for _grp in [us_tickers, si_tickers, hk_tickers]:
            _fetch_group(_grp)
        app.logger.info(f"[compute] prices after fetch: {prices}")

        # Fallback: use stored current_price
        for eq in equities:
            if eq["ticker"] not in prices and eq.get("current_price"):
                prices[eq["ticker"]] = float(eq["current_price"])

        # ── Price date (what date did we get prices for) ───────────
        # price_date already set above from actual yfinance data

        # ── Pre-fetch FX rates ─────────────────────────────────────
        ccys = {(e.get("currency") or "USD").upper() for e in equities}
        ccys |= {(c.get("currency") or "USD").upper() for c in cryptos}
        ccys.discard("SGD"); ccys.discard("USD")
        if ccys:
            _batch_fx_rates([f"{c}SGD" for c in ccys])

        # ── Equity valuation ──────────────────────────────────────
        equity_detail = []
        equities_sgd  = 0.0
        for eq in equities:
            tk     = eq["ticker"]
            shares = float(eq["shares"] or 0)
            ccy    = (eq.get("currency") or "USD").upper()
            px     = prices.get(tk, float(eq.get("current_price") or 0))
            if ccy == "SGD":
                val_sgd = px * shares
            elif ccy == "USD":
                val_sgd = px * shares * sgd_per_usd
            else:
                fx  = get_fx_rate(ccy, "SGD")
                val_sgd = px * shares * fx
            equities_sgd += val_sgd
            equity_detail.append({
                "ticker": tk, "shares": shares, "currency": ccy,
                "price_native": round(px, 4),
                "value_sgd": round(val_sgd, 2),
                "purchase_date": eq.get("purchase_date"),
            })

        # ── Crypto valuation ──────────────────────────────────────
        crypto_detail = []
        crypto_sgd    = 0.0
        crypto_syms   = list({c["symbol"] for c in cryptos})
        crypto_prices = {}
        if crypto_syms:
            try:
                with _suppress():
                    craw = yf.download(
                        [f"{s}-USD" for s in crypto_syms],
                        start=price_start, end=price_end,
                        auto_adjust=True, progress=False, threads=True
                    )
                if not craw.empty:
                    cclose = craw["Close"] if "Close" in craw else craw
                    for s in crypto_syms:
                        col = f"{s}-USD"
                        if len(crypto_syms) == 1:
                            if not cclose.empty:
                                crypto_prices[s] = float(cclose.iloc[-1])
                        elif col in cclose.columns and not cclose[col].dropna().empty:
                            crypto_prices[s] = float(cclose[col].dropna().iloc[-1])
            except Exception:
                pass
        for cr in cryptos:
            sym = cr["symbol"]
            qty = float(cr["quantity"] or 0)
            px  = crypto_prices.get(sym, float(cr.get("current_price") or 0))
            ccy = (cr.get("currency") or "USD").upper()
            val_usd = px * qty
            val_sgd = to_sgd(val_usd, "USD") if ccy == "USD" else to_sgd(val_usd, ccy)
            crypto_sgd += val_sgd
            crypto_detail.append({"symbol": sym, "qty": qty,
                                   "price_usd": round(px, 4), "value_sgd": round(val_sgd, 2)})

        # ── Cash ─────────────────────────────────────────────────
        cash_sgd = sum(to_sgd(r["amount"], r["currency"]) for r in cash_rows)

        # ── Trading total ─────────────────────────────────────────
        trading_total_sgd = equities_sgd + crypto_sgd + cash_sgd

        # ── Finance (networth_items) ──────────────────────────────
        nw_items = list_networth_items()
        fin_by_cat: dict[str, float] = {}
        for item in nw_items:
            cat = item.get("category", "other")
            amt = to_sgd(item.get("amount", 0), item.get("currency", "SGD"))
            fin_by_cat[cat] = fin_by_cat.get(cat, 0) + amt

        finance_assets_sgd = sum(v for k, v in fin_by_cat.items() if k != "liability")
        finance_liabs_sgd  = fin_by_cat.get("liability", 0)

        total_assets_sgd = finance_assets_sgd + trading_total_sgd
        net_worth_sgd    = total_assets_sgd - finance_liabs_sgd

        # ── Options stats ─────────────────────────────────────────
        with get_conn() as conn:
            open_count = conn.execute(
                "SELECT COUNT(*) FROM trades WHERE status='OPEN'"
            ).fetchone()[0]
            closed = conn.execute(
                "SELECT pnl FROM trades WHERE status='CLOSED' AND pnl IS NOT NULL"
            ).fetchall()
        realized_pnl_usd = sum(r[0] for r in closed)
        realized_pnl_sgd = realized_pnl_usd * sgd_per_usd
        wins = sum(1 for r in closed if r[0] > 0)
        win_rate = round(wins / len(closed) * 100, 1) if closed else 0.0

        # ── Build breakdown_json ──────────────────────────────────
        breakdown = {
            "finances": fin_by_cat,
            "trading":  {
                "equities_sgd": round(equities_sgd, 2),
                "crypto_sgd":   round(crypto_sgd, 2),
                "cash_sgd":     round(cash_sgd, 2),
                "total_sgd":    round(trading_total_sgd, 2),
            },
            "options":  {
                "realized_pnl_usd": round(realized_pnl_usd, 2),
                "realized_pnl_sgd": round(realized_pnl_sgd, 2),
                "open_count":       open_count,
            },
            "price_date":     price_date,
            "fx_usdsgd":      sgd_per_usd,      # 1 USD = X SGD
            "fx_sgdusd":      fetched_fx_sgd,   # 1 SGD = X USD (display)
            "equity_detail": equity_detail,
            "crypto_detail": crypto_detail,
        }

        # ── Save snapshot ─────────────────────────────────────────
        snap_data = {
            "month":            month,
            "portfolio_value":  round(trading_total_sgd, 2),
            "equities_value":   round(equities_sgd, 2),
            "crypto_value":     round(crypto_sgd, 2),
            "options_pnl":      round(realized_pnl_sgd, 2),
            "options_open":     open_count,
            "win_rate":         win_rate,
            "net_worth":        round(net_worth_sgd, 2),
            "total_assets":     round(total_assets_sgd, 2),
            "total_liabilities": round(finance_liabs_sgd, 2),
            "breakdown_json":   json.dumps(breakdown),
        }
        saved = save_monthly_snapshot(snap_data)
        return jsonify({"snapshot": saved, "breakdown": breakdown})

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/snapshots", methods=["GET"])
def get_snapshots():
    from database import list_monthly_snapshots
    return jsonify(list_monthly_snapshots())

@app.route("/api/snapshots", methods=["POST"])
def save_snapshot():
    from database import save_monthly_snapshot
    data = request.json or {}
    if not data.get("month"):
        return jsonify({"error": "month required"}), 400
    return jsonify(save_monthly_snapshot(data))

@app.route("/api/snapshots/<month>", methods=["DELETE"])
def delete_snapshot(month):
    from database import delete_monthly_snapshot
    deleted = delete_monthly_snapshot(month)
    return jsonify({"ok": deleted})


@app.route("/api/snapshots/<month>/breakdown.xlsx", methods=["GET"])
def snapshot_breakdown_xlsx(month):
    """
    Download an Excel reconciliation of the trading portfolio for a given snapshot month.
    Columns: Ticker/Item, Qty, Currency, Price (EOM), Local Value, SGD Value
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter
    from database import get_monthly_snapshot

    snap = get_monthly_snapshot(month)
    if not snap:
        abort(404)

    bd = {}
    try:
        bd = json.loads(snap.get("breakdown_json") or "{}")
    except Exception:
        pass

    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"NW Breakdown {month}"

    FONT     = "Calibri"
    C_DARK   = "FF0A0E1A"
    C_NAV    = "FF0D1B2E"
    C_ACCENT = "FF00D4FF"
    C_GREEN  = "FF00C882"
    C_WHITE  = "FFFFFFFF"
    C_MUTED  = "FF8899AA"
    C_ROW_ALT= "FF0F1E30"
    NUM_FMT  = "#,##0.00"
    SGD_FMT  = '"S$"#,##0.00'
    INT_FMT  = "#,##0"

    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col) if hex_col else None

    def hdr(bold=True, color=C_WHITE, size=10):
        return Font(name=FONT, bold=bold, size=size, color=color)

    cell_border = Border(bottom=Side(style="thin", color="FF1A2A40"))

    def set_cell(ws, r, c, val, font=None, bg=None, align="left", fmt=None, border=None):
        cell = ws.cell(row=r, column=c, value=val)
        if font:   cell.font      = font
        if bg:     cell.fill      = fill(bg)
        if fmt:    cell.number_format = fmt
        if border: cell.border    = border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        return cell

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    row = 1
    ws.merge_cells("A1:G1")
    set_cell(ws, 1, 1, f"Net Worth Breakdown -- {month}",
             font=Font(name=FONT, bold=True, size=13, color=C_ACCENT), bg=C_NAV)
    ws.row_dimensions[1].height = 22
    row = 2
    ws.merge_cells("A2:G2")
    set_cell(ws, 2, 1,
             f"Prices as of: {bd.get('price_date', month)}   |   FX: 1 SGD = {bd.get('fx_usd_per_sgd',0.74)} USD",
             font=Font(name=FONT, size=8, color=C_MUTED), bg=C_NAV)
    ws.row_dimensions[2].height = 14
    row = 3

    hdr_row = row
    headers = ["Ticker / Item", "Qty", "CCY", "Buy Date", "Price (EOM)", "Local Value", "SGD Value"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, row, i, h, font=hdr(True, C_DARK, 9), bg=C_ACCENT,
                 align="center" if i > 1 else "left")
    ws.row_dimensions[row].height = 16
    row += 1

    def section_header(ws, r, label):
        ws.merge_cells(f"A{r}:G{r}")
        set_cell(ws, r, 1, label, font=hdr(True, C_ACCENT, 9), bg=C_NAV)
        ws.row_dimensions[r].height = 14

    def data_row(ws, r, ticker, qty, ccy, buy_date, price, local_val, sgd_val, alt=False):
        bg = C_ROW_ALT if alt else None
        br = cell_border
        set_cell(ws, r, 1, ticker,    font=hdr(False, C_WHITE, 9),  bg=bg, border=br)
        set_cell(ws, r, 2, qty,       font=hdr(False, C_WHITE, 9),  bg=bg, align="right", fmt=INT_FMT, border=br)
        set_cell(ws, r, 3, ccy,       font=hdr(False, C_MUTED, 9),  bg=bg, align="center", border=br)
        set_cell(ws, r, 4, buy_date,  font=hdr(False, C_MUTED, 9),  bg=bg, align="center", border=br)
        set_cell(ws, r, 5, price,     font=hdr(False, C_WHITE, 9),  bg=bg, align="right", fmt=NUM_FMT, border=br)
        set_cell(ws, r, 6, local_val, font=hdr(False, C_WHITE, 9),  bg=bg, align="right", fmt=SGD_FMT, border=br)
        set_cell(ws, r, 7, sgd_val,   font=hdr(False, C_GREEN, 9),  bg=bg, align="right", fmt=SGD_FMT, border=br)

    def total_row(ws, r, label, first_data_row, last_data_row):
        ws.merge_cells(f"A{r}:F{r}")
        set_cell(ws, r, 1, label, font=hdr(True, C_WHITE, 9), bg=C_NAV)
        total_cell = ws.cell(row=r, column=7,
                             value=f"=SUM(G{first_data_row}:G{last_data_row})")
        total_cell.font         = Font(name=FONT, bold=True, size=9, color=C_GREEN)
        total_cell.fill         = fill(C_NAV)
        total_cell.number_format = SGD_FMT
        total_cell.alignment    = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 14
        return total_cell

    section_header(ws, row, "  EQUITIES"); row += 1
    eq_first = row
    eq_detail = bd.get("equity_detail", [])
    if eq_detail:
        for i, e in enumerate(eq_detail):
            data_row(ws, row, e.get("ticker",""), e.get("shares",""),
                     e.get("currency",""), e.get("purchase_date",""),
                     e.get("price_native", 0),
                     e.get("price_native",0)*e.get("shares",0),
                     e.get("value_sgd", 0), alt=bool(i%2))
            row += 1
    else:
        data_row(ws, row, "(detail unavailable - re-capture snapshot)",
                 "", "", "", "", "", snap.get("equities_value", 0))
        row += 1
    eq_last = row - 1
    eq_total_cell = total_row(ws, row, "Equities Total", eq_first, eq_last)
    row += 2

    section_header(ws, row, "  CRYPTO"); row += 1
    cr_first = row
    cr_detail = bd.get("crypto_detail", [])
    if cr_detail:
        for i, c in enumerate(cr_detail):
            data_row(ws, row, c.get("symbol",""), c.get("qty",""),
                     "USD", "", c.get("price_usd",0),
                     c.get("price_usd",0)*c.get("qty",0),
                     c.get("value_sgd",0), alt=bool(i%2))
            row += 1
    else:
        data_row(ws, row, "(detail unavailable - re-capture snapshot)",
                 "", "", "", "", "", snap.get("crypto_value", 0))
        row += 1
    cr_last = row - 1
    cr_total_cell = total_row(ws, row, "Crypto Total", cr_first, cr_last)
    row += 2

    trd = bd.get("trading", {})
    cash_sgd = trd.get("cash_sgd")
    if cash_sgd is None:
        cash_sgd = (snap.get("portfolio_value", 0)
                    - snap.get("equities_value", 0)
                    - snap.get("crypto_value", 0))
    section_header(ws, row, "  TRADING CASH"); row += 1
    cash_row = row
    data_row(ws, row, "Trading accounts (aggregate)", "", "SGD", "", "", "", cash_sgd)
    row += 1
    cash_total_cell = total_row(ws, row, "Cash Total", cash_row, cash_row)
    row += 2

    ws.merge_cells(f"A{row}:F{row}")
    set_cell(ws, row, 1, "TRADING PORTFOLIO TOTAL",
             font=Font(name=FONT, bold=True, size=10, color=C_DARK), bg=C_ACCENT,
             border=Border(top=Side(style="medium", color=C_DARK)))
    grand_formula = (f"={eq_total_cell.coordinate}"
                     f"+{cr_total_cell.coordinate}"
                     f"+{cash_total_cell.coordinate}")
    gc = ws.cell(row=row, column=7, value=grand_formula)
    gc.font  = Font(name=FONT, bold=True, size=10, color=C_DARK)
    gc.fill  = fill(C_ACCENT)
    gc.number_format = SGD_FMT
    gc.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18
    grand_row = row
    row += 1

    set_cell(ws, row, 1, "Stored in snapshot",
             font=hdr(False, C_MUTED, 8), bg=C_DARK)
    set_cell(ws, row, 7, snap.get("portfolio_value", 0),
             font=hdr(False, C_MUTED, 8), bg=C_DARK, align="right", fmt=SGD_FMT)
    row += 1
    set_cell(ws, row, 1, "Rounding difference",
             font=hdr(False, C_MUTED, 8), bg=C_DARK)
    set_cell(ws, row, 7, f"=G{grand_row}-G{grand_row+1}",
             font=hdr(False, C_MUTED, 8), bg=C_DARK, align="right", fmt=SGD_FMT)
    row += 2

    fin = bd.get("finances", {})
    if fin:
        section_header(ws, row, "  FINANCES (networth_items)"); row += 1
        fin_first = row
        CAT_LABELS = {
            "cash": "Cash in Bank", "cpf_srs": "CPF / SRS",
            "property": "Property", "insurance": "Insurance",
            "investment": "Other Investments", "other": "Other Assets",
            "liability": "Liabilities",
        }
        for i, (cat, val) in enumerate(fin.items()):
            label = CAT_LABELS.get(cat, cat.title())
            color = "FFFF6B6B" if cat == "liability" else C_WHITE
            set_cell(ws, row, 1, label, font=hdr(False, color, 9),
                     bg=C_ROW_ALT if i%2 else None)
            set_cell(ws, row, 7, val, font=hdr(False, color, 9),
                     bg=C_ROW_ALT if i%2 else None,
                     align="right", fmt=SGD_FMT, border=cell_border)
            row += 1
        total_row(ws, row, "Finance Total (assets excl. liabilities)", fin_first, row-1)
        row += 1

    ws.freeze_panes = f"A{hdr_row+1}"
    ws.sheet_view.showGridLines = False

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=nw_breakdown_{month}.xlsx"}
    )


# -- Main --
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=not DEMO_MODE, port=port)